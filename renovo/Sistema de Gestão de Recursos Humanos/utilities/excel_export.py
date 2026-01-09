"""
Módulo de Exportação Excel - Sistema de Gestão de RH
RENOVO Montagens Industriais
"""

import os
from datetime import datetime
from typing import List, Dict, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _criar_estilo_header():
    """Retorna estilos para cabeçalhos."""
    return {
        'font': Font(bold=True, color="FFFFFF", size=10),
        'fill': PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
        'border': Border(
            left=Side(style='thin', color='bdc3c7'),
            right=Side(style='thin', color='bdc3c7'),
            top=Side(style='thin', color='bdc3c7'),
            bottom=Side(style='thin', color='bdc3c7')
        )
    }


def _criar_estilo_celula():
    """Retorna estilo para células de dados."""
    return {
        'border': Border(
            left=Side(style='thin', color='bdc3c7'),
            right=Side(style='thin', color='bdc3c7'),
            top=Side(style='thin', color='bdc3c7'),
            bottom=Side(style='thin', color='bdc3c7')
        )
    }


def _formatar_data(data_str: str) -> str:
    """Formata data de YYYY-MM-DD para DD/MM/YYYY."""
    if not data_str:
        return ""
    try:
        data_obj = datetime.strptime(str(data_str), '%Y-%m-%d')
        return data_obj.strftime('%d/%m/%Y')
    except:
        return str(data_str)


def _formatar_moeda(valor) -> str:
    """Formata valor para moeda brasileira."""
    if not valor:
        return ""
    try:
        return f"R$ {float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return str(valor)


def _aplicar_header(ws, headers: List[tuple], row: int = 1):
    """Aplica cabeçalhos com estilo em uma planilha."""
    estilos = _criar_estilo_header()

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = estilos['font']
        cell.fill = estilos['fill']
        cell.alignment = estilos['alignment']
        cell.border = estilos['border']
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return row + 1


def _aplicar_dados(ws, dados: List[List], start_row: int):
    """Aplica dados com estilo em uma planilha."""
    estilos = _criar_estilo_celula()

    for row_idx, linha in enumerate(dados, start_row):
        for col_idx, valor in enumerate(linha, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = estilos['border']


def exportar_completo_excel(
    colaboradores_ativos: List[Dict],
    colaboradores_inativos: List[Dict],
    contratos: List[Dict],
    ferias: List[Dict],
    dependentes: List[Dict],
    blocklist: List[Dict],
    documentos_pendentes: List[Dict],
    output_path: str,
    empresa_nome: str = "Todas"
) -> str:
    """
    Exporta todos os dados do sistema para um único arquivo Excel com múltiplas abas.

    Abas:
    1. Colaboradores Ativos
    2. Colaboradores Inativos
    3. Contratos Experiência
    4. Férias
    5. Dependentes
    6. Block-List
    7. Documentos Pendentes
    """

    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")

    wb = openpyxl.Workbook()

    # Remove a aba padrão
    wb.remove(wb.active)

    # =========================================================================
    # ABA 1: Colaboradores Ativos
    # =========================================================================
    ws_ativos = wb.create_sheet("Colaboradores Ativos")

    headers_colab = [
        ("ID", 6),
        ("Nome Completo", 35),
        ("CPF", 15),
        ("Data Nascimento", 14),
        ("Grau Instrução", 18),
        ("Curso/Formação", 25),
        ("Função", 20),
        ("Departamento", 18),
        ("Data Admissão", 14),
        ("Salário", 14),
        ("Telefone", 15),
        ("Celular", 15),
        ("E-mail", 30),
        ("Cidade", 18),
        ("UF", 5),
        ("Empresa", 30),
        ("Data Exame (ASO)", 15),
        ("Tipo Exames", 20),
        ("Médico", 25),
        ("CRM", 12),
        ("Empresa Anterior", 25),
        ("Admissão Anterior", 14),
        ("Saída Anterior", 14),
    ]

    next_row = _aplicar_header(ws_ativos, headers_colab)

    dados_ativos = []
    for c in colaboradores_ativos:
        dados_ativos.append([
            c.get('id', ''),
            c.get('nome_completo', ''),
            c.get('cpf', ''),
            _formatar_data(c.get('data_nascimento', '')),
            c.get('grau_instrucao', ''),
            c.get('curso_formacao', ''),
            c.get('funcao', ''),
            c.get('departamento', ''),
            _formatar_data(c.get('data_admissao', '')),
            _formatar_moeda(c.get('salario', '')),
            c.get('telefone', ''),
            c.get('celular', ''),
            c.get('email', ''),
            c.get('cidade', ''),
            c.get('uf_endereco', ''),
            c.get('empresa_nome', ''),
            _formatar_data(c.get('data_exame_medico', '')),
            c.get('tipo_exames', ''),
            c.get('nome_medico', ''),
            f"{c.get('crm', '')}/{c.get('uf_crm', '')}" if c.get('crm') else '',
            c.get('empresa_ultimo_emprego', ''),
            _formatar_data(c.get('data_admissao_ultimo', '')),
            _formatar_data(c.get('data_saida_ultimo', '')),
        ])

    _aplicar_dados(ws_ativos, dados_ativos, next_row)
    ws_ativos.freeze_panes = 'A2'
    ws_ativos.auto_filter.ref = f"A1:{get_column_letter(len(headers_colab))}1"

    # =========================================================================
    # ABA 2: Colaboradores Inativos
    # =========================================================================
    ws_inativos = wb.create_sheet("Colaboradores Inativos")

    headers_inativos = [
        ("ID", 6),
        ("Nome Completo", 35),
        ("CPF", 15),
        ("Função", 20),
        ("Data Admissão", 14),
        ("Data Inativação", 14),
        ("Motivo Inativação", 25),
        ("Submotivo", 25),
        ("Empresa", 30),
        ("Últ. Exame (ASO)", 15),
        ("Tipo Exames", 20),
    ]

    next_row = _aplicar_header(ws_inativos, headers_inativos)

    dados_inativos = []
    for c in colaboradores_inativos:
        dados_inativos.append([
            c.get('id', ''),
            c.get('nome_completo', ''),
            c.get('cpf', ''),
            c.get('funcao', ''),
            _formatar_data(c.get('data_admissao', '')),
            _formatar_data(c.get('data_inativacao', '') or c.get('data_desligamento', '')),
            c.get('motivo_inativacao', '') or c.get('motivo_desligamento', ''),
            c.get('submotivo_inativacao', ''),
            c.get('empresa_nome', ''),
            _formatar_data(c.get('data_exame_medico', '')),
            c.get('tipo_exames', ''),
        ])

    _aplicar_dados(ws_inativos, dados_inativos, next_row)
    ws_inativos.freeze_panes = 'A2'

    # =========================================================================
    # ABA 3: Contratos de Experiência
    # =========================================================================
    ws_contratos = wb.create_sheet("Contratos Experiência")

    headers_contratos = [
        ("Colaborador", 35),
        ("CPF", 15),
        ("Função", 20),
        ("Data Início", 12),
        ("Prazo Inicial", 12),
        ("Fim Inicial", 12),
        ("Prorrogação", 12),
        ("Fim Prorrogação", 14),
        ("Status", 12),
        ("Empresa", 30),
    ]

    next_row = _aplicar_header(ws_contratos, headers_contratos)

    dados_contratos = []
    for c in contratos:
        dados_contratos.append([
            c.get('nome_completo', ''),
            c.get('cpf', ''),
            c.get('funcao', ''),
            _formatar_data(c.get('data_inicio', '')),
            f"{c.get('prazo_inicial', '')} dias" if c.get('prazo_inicial') else '',
            _formatar_data(c.get('data_fim_inicial', '')),
            f"{c.get('prorrogacao', '')} dias" if c.get('prorrogacao') else '',
            _formatar_data(c.get('data_fim_prorrogacao', '')),
            c.get('status', ''),
            c.get('empresa_nome', ''),
        ])

    _aplicar_dados(ws_contratos, dados_contratos, next_row)
    ws_contratos.freeze_panes = 'A2'

    # =========================================================================
    # ABA 4: Férias
    # =========================================================================
    ws_ferias = wb.create_sheet("Férias")

    headers_ferias = [
        ("Colaborador", 35),
        ("CPF", 15),
        ("Período Aquisitivo Início", 20),
        ("Período Aquisitivo Fim", 18),
        ("Limite Concessivo", 16),
        ("Dias Direito", 12),
        ("Dias Gozados", 12),
        ("Dias Vendidos", 12),
        ("Status", 12),
        ("Empresa", 30),
    ]

    next_row = _aplicar_header(ws_ferias, headers_ferias)

    dados_ferias = []
    for f in ferias:
        dados_ferias.append([
            f.get('nome_completo', ''),
            f.get('cpf', ''),
            _formatar_data(f.get('periodo_aquisitivo_inicio', '')),
            _formatar_data(f.get('periodo_aquisitivo_fim', '')),
            _formatar_data(f.get('periodo_concessivo_limite', '')),
            f.get('dias_direito', 30),
            f.get('dias_gozados', 0),
            f.get('dias_vendidos', 0),
            f.get('status', ''),
            f.get('empresa_nome', ''),
        ])

    _aplicar_dados(ws_ferias, dados_ferias, next_row)
    ws_ferias.freeze_panes = 'A2'

    # =========================================================================
    # ABA 5: Dependentes
    # =========================================================================
    ws_dependentes = wb.create_sheet("Dependentes")

    headers_dep = [
        ("Colaborador", 35),
        ("CPF Colaborador", 15),
        ("Nome Dependente", 35),
        ("Parentesco", 15),
        ("Data Nascimento", 14),
        ("CPF Dependente", 15),
    ]

    next_row = _aplicar_header(ws_dependentes, headers_dep)

    dados_dep = []
    for d in dependentes:
        dados_dep.append([
            d.get('colaborador_nome', ''),
            d.get('colaborador_cpf', ''),
            d.get('nome', ''),
            d.get('parentesco', ''),
            _formatar_data(d.get('data_nascimento', '')),
            d.get('cpf', ''),
        ])

    _aplicar_dados(ws_dependentes, dados_dep, next_row)
    ws_dependentes.freeze_panes = 'A2'

    # =========================================================================
    # ABA 6: Block-List
    # =========================================================================
    ws_blocklist = wb.create_sheet("Block-List")

    headers_block = [
        ("Nome", 35),
        ("CPF", 15),
        ("Empresa", 30),
        ("Data Admissão", 14),
        ("Data Desligamento", 14),
        ("Motivo Desligamento", 30),
        ("Pode Recontratar", 15),
        ("Observações", 40),
    ]

    next_row = _aplicar_header(ws_blocklist, headers_block)

    dados_block = []
    for b in blocklist:
        dados_block.append([
            b.get('nome', ''),
            b.get('cpf', ''),
            b.get('empresa_nome', ''),
            _formatar_data(b.get('data_admissao', '')),
            _formatar_data(b.get('data_desligamento', '')),
            b.get('motivo_desligamento', ''),
            'Sim' if b.get('pode_recontratar', 1) == 1 else 'Não',
            b.get('observacoes', ''),
        ])

    _aplicar_dados(ws_blocklist, dados_block, next_row)
    ws_blocklist.freeze_panes = 'A2'

    # =========================================================================
    # ABA 7: Documentos Pendentes
    # =========================================================================
    ws_docs = wb.create_sheet("Documentos Pendentes")

    headers_docs = [
        ("Colaborador", 35),
        ("CPF", 15),
        ("Documento Pendente", 40),
        ("Tipo", 15),
    ]

    next_row = _aplicar_header(ws_docs, headers_docs)

    dados_docs = []
    for d in documentos_pendentes:
        dados_docs.append([
            d.get('colaborador_nome', ''),
            d.get('colaborador_cpf', ''),
            d.get('documento', ''),
            'Obrigatório' if d.get('obrigatorio', True) else 'Opcional',
        ])

    _aplicar_dados(ws_docs, dados_docs, next_row)
    ws_docs.freeze_panes = 'A2'

    # =========================================================================
    # Salvar arquivo
    # =========================================================================
    wb.save(output_path)

    return output_path


# Manter funções antigas para compatibilidade (caso sejam usadas em outro lugar)
def exportar_colaboradores_excel(colaboradores: List[Dict], output_dir: str = "exports",
                                  filename: str = None) -> str:
    """Exporta lista de colaboradores para Excel. (Função legada)"""

    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não está instalado. Execute: pip install openpyxl")

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"colaboradores_{timestamp}.xlsx"

    output_path = os.path.join(output_dir, filename)

    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colaboradores"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_border = Border(
        left=Side(style='thin', color='bdc3c7'),
        right=Side(style='thin', color='bdc3c7'),
        top=Side(style='thin', color='bdc3c7'),
        bottom=Side(style='thin', color='bdc3c7')
    )

    # Definir colunas
    colunas = [
        ("ID", "id", 8),
        ("Nome Completo", "nome_completo", 35),
        ("CPF", "cpf", 15),
        ("Data Nascimento", "data_nascimento", 15),
        ("Empresa", "empresa_nome", 30),
        ("Função", "funcao", 20),
        ("Departamento", "departamento", 18),
        ("Data Admissão", "data_admissao", 15),
        ("Salário", "salario", 12),
        ("Status", "status", 12),
        ("Telefone", "telefone", 15),
        ("Celular", "celular", 15),
        ("E-mail", "email", 30),
        ("Cidade", "cidade", 18),
        ("UF", "uf_endereco", 5),
    ]

    # Escrever cabeçalhos
    for col_idx, (header, _, width) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Escrever dados
    for row_idx, colab in enumerate(colaboradores, 2):
        for col_idx, (_, field, _) in enumerate(colunas, 1):
            valor = colab.get(field, '')

            # Formatar valores especiais
            if field == 'salario' and valor:
                try:
                    valor = float(valor)
                except:
                    pass
            elif field in ['data_nascimento', 'data_admissao'] and valor:
                try:
                    if isinstance(valor, str):
                        valor = datetime.strptime(valor, '%Y-%m-%d').strftime('%d/%m/%Y')
                except:
                    pass

            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = cell_border

            # Formato para salário
            if field == 'salario':
                cell.number_format = 'R$ #,##0.00'

    # Congelar primeira linha
    ws.freeze_panes = 'A2'

    # Adicionar filtro automático
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}1"

    # Salvar
    wb.save(output_path)

    return output_path


def exportar_aniversariantes_excel(aniversariantes: List[Dict], mes: int,
                                    output_dir: str = "exports") -> str:
    """Exporta lista de aniversariantes para Excel. (Função legada)"""

    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não está instalado")

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"aniversariantes_{meses[mes]}_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Aniversariantes - {meses[mes]}"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")

    # Título
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"ANIVERSARIANTES DE {meses[mes].upper()}"
    title_cell.font = Font(bold=True, size=14, color="1a5276")
    title_cell.alignment = Alignment(horizontal="center")

    # Cabeçalhos
    headers = ["Nome", "Data Nascimento", "Dia", "Função", "Empresa"]
    widths = [35, 18, 8, 25, 30]

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Dados
    for row_idx, aniv in enumerate(aniversariantes, 4):
        ws.cell(row=row_idx, column=1, value=aniv.get('nome_completo', ''))

        data_nasc = aniv.get('data_nascimento', '')
        if data_nasc:
            try:
                data_obj = datetime.strptime(str(data_nasc), '%Y-%m-%d')
                ws.cell(row=row_idx, column=2, value=data_obj.strftime('%d/%m/%Y'))
                ws.cell(row=row_idx, column=3, value=data_obj.day)
            except:
                ws.cell(row=row_idx, column=2, value=data_nasc)

        ws.cell(row=row_idx, column=4, value=aniv.get('funcao', ''))
        ws.cell(row=row_idx, column=5, value=aniv.get('empresa_nome', ''))

    wb.save(output_path)
    return output_path


def exportar_contratos_vencendo_excel(contratos: List[Dict], output_dir: str = "exports") -> str:
    """Exporta lista de contratos de experiência vencendo para Excel. (Função legada)"""

    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não está instalado")

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"contratos_vencendo_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contratos Vencendo"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="c0392b", end_color="c0392b", fill_type="solid")

    headers = ["Colaborador", "Empresa", "Função", "Início", "Fim Inicial",
               "Prorrogação", "Fim Prorrogação", "Status"]
    widths = [30, 25, 20, 12, 12, 12, 15, 12]

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, contrato in enumerate(contratos, 2):
        ws.cell(row=row_idx, column=1, value=contrato.get('nome_completo', ''))
        ws.cell(row=row_idx, column=2, value=contrato.get('empresa_nome', ''))
        ws.cell(row=row_idx, column=3, value=contrato.get('funcao', ''))
        ws.cell(row=row_idx, column=4, value=contrato.get('data_inicio', ''))
        ws.cell(row=row_idx, column=5, value=contrato.get('data_fim_inicial', ''))
        ws.cell(row=row_idx, column=6, value=contrato.get('prorrogacao', ''))
        ws.cell(row=row_idx, column=7, value=contrato.get('data_fim_prorrogacao', ''))
        ws.cell(row=row_idx, column=8, value=contrato.get('status', ''))

    wb.save(output_path)
    return output_path


def exportar_ferias_vencendo_excel(ferias: List[Dict], output_dir: str = "exports") -> str:
    """Exporta lista de férias com período concessivo vencendo. (Função legada)"""

    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl não está instalado")

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ferias_vencendo_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Férias Vencendo"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")

    headers = ["Colaborador", "Empresa", "Período Aquisitivo", "Limite Concessivo",
               "Dias Direito", "Dias Gozados", "Dias Vendidos", "Status"]
    widths = [30, 25, 25, 15, 12, 12, 12, 12]

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, f in enumerate(ferias, 2):
        ws.cell(row=row_idx, column=1, value=f.get('nome_completo', ''))
        ws.cell(row=row_idx, column=2, value=f.get('empresa_nome', ''))

        periodo = f"{f.get('periodo_aquisitivo_inicio', '')} a {f.get('periodo_aquisitivo_fim', '')}"
        ws.cell(row=row_idx, column=3, value=periodo)
        ws.cell(row=row_idx, column=4, value=f.get('periodo_concessivo_limite', ''))
        ws.cell(row=row_idx, column=5, value=f.get('dias_direito', 30))
        ws.cell(row=row_idx, column=6, value=f.get('dias_gozados', 0))
        ws.cell(row=row_idx, column=7, value=f.get('dias_vendidos', 0))
        ws.cell(row=row_idx, column=8, value=f.get('status', ''))

    wb.save(output_path)
    return output_path
