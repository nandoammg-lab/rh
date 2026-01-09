"""
Sistema de Gestão de RH - RENOVO Montagens Industriais
Aplicação Principal

Para executar: flet run app.py
Para web: flet run app.py --web --port 8080
"""

import sys
import os
import multiprocessing
import argparse
import subprocess

# CRÍTICO: Corrigir sys.stdin/stdout/stderr para PyInstaller --noconsole
# Isso evita erros do uvicorn ao tentar acessar isatty() em None
if getattr(sys, 'frozen', False):
    # Executável compilado com --noconsole: streams podem ser None
    if sys.stdout is None:
        sys.stdout = open(os.devnull, 'w')
    if sys.stderr is None:
        sys.stderr = open(os.devnull, 'w')
    if sys.stdin is None:
        sys.stdin = open(os.devnull, 'r')

# CRÍTICO: freeze_support DEVE ser chamado no início para PyInstaller --onefile
multiprocessing.freeze_support()

import flet as ft
import json
from datetime import datetime, timedelta
import time
import asyncio
import shutil


def get_base_path():
    """
    Retorna o caminho base do executável ou script.
    Necessário para PyInstaller --onefile funcionar corretamente.
    """
    if getattr(sys, 'frozen', False):
        # Executando como executável PyInstaller
        return os.path.dirname(sys.executable)
    else:
        # Executando como script Python
        return os.path.dirname(os.path.abspath(__file__))


def parse_args():
    """Parse argumentos da linha de comando"""
    parser = argparse.ArgumentParser(description='Sistema de Gestão de RH')
    parser.add_argument('--usuario', type=str, default='', help='Login do usuário')
    parser.add_argument('--nome', type=str, default='', help='Nome do usuário')
    parser.add_argument('--cargo', type=str, default='', help='Cargo do usuário')
    parser.add_argument('--erp_path', type=str, default='', help='Caminho do ERP')
    args, unknown = parser.parse_known_args()
    return args


# Dados do usuário vindos do ERP
ERP_USER_DATA = parse_args()


from utilities import database as db
from utilities.main import (
    criar_alertas_widget, criar_campo_view, criar_secao, FichaColaborador,
    formatar_cpf, formatar_data_br, formatar_moeda,
    COR_PRIMARIA, COR_SECUNDARIA, COR_SUCESSO, COR_ALERTA, COR_ERRO, COR_FUNDO, COR_CINZA_CLARO
)
from utilities.formulario_cadastro import FormularioCadastro
from utilities.excel_export import (
    exportar_colaboradores_excel, exportar_aniversariantes_excel,
    exportar_contratos_vencendo_excel, exportar_ferias_vencendo_excel
)
from utilities.pdf_generator import gerar_ficha_registro_pdf
from utilities.dashboard import DashboardView


class SistemaRH:
    """Classe principal do Sistema de Gestão de RH."""
    
    # Arquivo para persistir dados temporários do formulário
    ARQUIVO_DADOS_TEMP = "dados_formulario_temp.json"

    def __init__(self, page: ft.Page):
        self.page = page
        self.configurar_pagina()
        self.view_atual = "lista"

        # Armazenamento temporário dos dados do formulário de cadastro
        self.dados_formulario_temp = {}
        self.dependentes_temp = []
        self.foto_path_temp = None
        self._formulario_atual = None  # Referência ao formulário de cadastro atual

        # Carregar dados temporários persistidos (se existirem)
        self._carregar_dados_temp_arquivo()

        self.container_principal = ft.Container(expand=True, alignment=ft.alignment.top_left)
        self.construir_interface()
    
    def configurar_pagina(self):
        self.page.title = "Sistema de Gestão de RH - RENOVO"
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.bgcolor = COR_FUNDO
        self.page.padding = 0
        self.page.scroll = None
        self.page.window.maximized = True

        # Configurar evento de fechamento para fazer backup automático
        self.page.window.prevent_close = True
        self.page.window.on_event = self._on_window_event

    def _on_window_event(self, e):
        """Trata eventos da janela, incluindo fechamento."""
        if e.data == "close":
            self._confirmar_saida()

    def _criar_avatar_foto(self, foto_path: str, cor_fallback=None, tamanho: int = 40) -> ft.Container:
        """
        Cria um widget de avatar com foto ou ícone de fallback.

        Args:
            foto_path: Caminho para a foto do colaborador
            cor_fallback: Cor de fundo quando não há foto (padrão: COR_SECUNDARIA)
            tamanho: Tamanho do avatar em pixels (padrão: 40)

        Returns:
            ft.Container: Widget do avatar
        """
        if cor_fallback is None:
            cor_fallback = COR_SECUNDARIA

        raio = tamanho // 2

        if foto_path and os.path.exists(foto_path):
            return ft.Container(
                content=ft.Image(src=foto_path, width=tamanho, height=tamanho, fit=ft.ImageFit.COVER, border_radius=raio),
                width=tamanho,
                height=tamanho,
                border_radius=raio,
                clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
            )
        else:
            return ft.Container(
                content=ft.Icon(ft.Icons.PERSON, size=tamanho * 0.6, color=ft.Colors.WHITE),
                width=tamanho,
                height=tamanho,
                bgcolor=cor_fallback,
                border_radius=raio,
                alignment=ft.alignment.center,
            )

    def _criar_dialogo_confirmacao(self, titulo: str, mensagem: str, on_confirmar,
                                    texto_confirmar: str = "Confirmar",
                                    texto_cancelar: str = "Cancelar",
                                    cor_confirmar=None) -> ft.AlertDialog:
        """
        Cria um diálogo de confirmação padronizado.

        Args:
            titulo: Título do diálogo
            mensagem: Mensagem de confirmação
            on_confirmar: Callback executado ao confirmar
            texto_confirmar: Texto do botão de confirmação
            texto_cancelar: Texto do botão de cancelar
            cor_confirmar: Cor do botão de confirmação (padrão: COR_PRIMARIA)

        Returns:
            ft.AlertDialog: Diálogo configurado
        """
        if cor_confirmar is None:
            cor_confirmar = COR_PRIMARIA

        dialog = ft.AlertDialog(modal=True)

        def fechar(_):
            dialog.open = False
            self.page.update()

        def confirmar(_):
            dialog.open = False
            self.page.update()
            on_confirmar()

        dialog.title = ft.Text(titulo)
        dialog.content = ft.Text(mensagem)
        dialog.actions = [
            ft.TextButton(texto_cancelar, on_click=fechar),
            ft.ElevatedButton(texto_confirmar, on_click=confirmar, bgcolor=cor_confirmar, color="white"),
        ]

        return dialog

    def _criar_card_stat(self, titulo: str, valor, icone, cor,
                         tamanho: str = "normal") -> ft.Container:
        """
        Cria um card de estatística padronizado.

        Args:
            titulo: Título do card
            valor: Valor numérico a exibir
            icone: Ícone do card
            cor: Cor do ícone/borda
            tamanho: "normal" ou "compacto"

        Returns:
            ft.Container: Card de estatística
        """
        if tamanho == "compacto":
            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(icone, size=18, color=cor),
                        ft.Text(titulo, size=10, color=ft.Colors.GREY_700),
                    ], spacing=5),
                    ft.Text(str(valor), size=20, weight=ft.FontWeight.BOLD, color=cor),
                ], spacing=3, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                padding=10,
                bgcolor="white",
                border_radius=8,
                border=ft.border.all(1, cor),
                width=100,
            )
        else:
            return ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(icone, size=24, color=cor),
                        ft.Text(titulo, size=12, color=ft.Colors.GREY_700),
                    ], spacing=5),
                    ft.Text(str(valor), size=28, weight=ft.FontWeight.BOLD, color=cor),
                ], spacing=5, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                padding=15,
                bgcolor="white",
                border_radius=8,
                border=ft.border.all(1, cor),
                width=140,
            )

    def _confirmar_saida(self):
        """Fecha o sistema, salvando dados temporários do formulário se necessário."""
        # Se estiver no cadastro de novo colaborador, salvar dados temporários
        if self.view_atual == "cadastro" and self._formulario_atual:
            try:
                dados_temp = self._formulario_atual._coletar_dados_temp()
                self._salvar_dados_temp(
                    dados_temp,
                    self._formulario_atual.dependentes_lista,
                    self._formulario_atual.foto_path
                )
            except Exception as e:
                db.registrar_log("sistema", "aviso", f"Erro ao salvar dados temporários ao fechar: {str(e)}")
        self.page.window.destroy()
    
    def construir_interface(self):
        self.menu_lateral = self._criar_menu_lateral()
        self.atualizar_view()

        self.layout_principal = ft.Row([
            self.menu_lateral,
            ft.VerticalDivider(width=1, color=COR_SECUNDARIA),
            ft.Container(
                content=self.container_principal,
                expand=True,
                padding=20,
            ),
        ], expand=True, spacing=0, vertical_alignment=ft.CrossAxisAlignment.START)

        self.page.add(self.layout_principal)
    
    def _criar_menu_lateral(self):
        # Caminho da logo
        logo_path = os.path.join(get_base_path(), "imagens", "Logomarca Renovo.png")

        # Obter usuário logado
        usuario = db.get_usuario_logado()
        nome_usuario = usuario.get('nome_completo', 'Usuário') if usuario else 'Usuário'
        nivel_acesso = usuario.get('nivel_acesso', 'operador') if usuario else 'operador'

        # Texto do nível de acesso formatado
        niveis_texto = {
            'administrador': 'Administrador',
            'operador': 'Operador',
            'visualizador': 'Visualizador'
        }
        nivel_texto = niveis_texto.get(nivel_acesso, 'Operador')

        # Itens do menu baseados nas permissões
        itens_menu = [
            self._item_menu("Colaboradores", ft.Icons.PEOPLE, "lista"),
        ]

        # Apenas operador e admin podem adicionar
        if db.usuario_pode('adicionar_colaborador'):
            itens_menu.append(self._item_menu("Novo Colaborador", ft.Icons.PERSON_ADD, "cadastro"))

        itens_menu.extend([
            ft.Divider(color=COR_SECUNDARIA),
            self._item_menu("Contratos Exp.", ft.Icons.SCHEDULE, "contratos"),
            self._item_menu("Férias", ft.Icons.BEACH_ACCESS, "ferias"),
            self._item_menu("Aniversariantes", ft.Icons.CAKE, "aniversariantes"),
            ft.Divider(color=COR_SECUNDARIA),
            self._item_menu("Banco de Talentos", ft.Icons.FOLDER_SHARED, "banco_talentos"),
            self._item_menu("Block-List", ft.Icons.BLOCK, "blocklist"),
            self._item_menu("Empresas", ft.Icons.BUSINESS, "empresas"),
        ])

        # Dashboard - apenas admin do ERP
        if ERP_USER_DATA.cargo and 'admin' in ERP_USER_DATA.cargo.lower():
            itens_menu.extend([
                ft.Divider(color=COR_SECUNDARIA),
                ft.Container(
                    content=ft.Text("ANÁLISES", size=10, color=ft.Colors.GREY_500, weight=ft.FontWeight.BOLD),
                    padding=ft.padding.only(left=15, top=5, bottom=5),
                ),
                self._item_menu("Dashboard", ft.Icons.ANALYTICS, "dashboard"),
            ])

        itens_menu.extend([
            ft.Divider(color=COR_SECUNDARIA),
            self._item_menu("Exportar Excel", ft.Icons.DOWNLOAD, "exportar"),
        ])

        # Backup - disponível para todos quando vem do ERP
        itens_menu.append(self._item_menu("Backup", ft.Icons.BACKUP, "backup"))

        itens_menu.extend([
            ft.Divider(color=COR_SECUNDARIA),
            self._item_menu("Log", ft.Icons.HISTORY, "log"),
        ])

        # Configurações removidas - o gerenciamento de usuários agora é feito pelo ERP

        itens_menu.append(self._item_menu("Sobre", ft.Icons.INFO, "sobre"))

        # Seção do ERP (se veio do ERP)
        erp_section = None
        if ERP_USER_DATA.nome:
            erp_section = ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.PERSON, size=14, color=COR_PRIMARIA),
                        ft.Text(
                            ERP_USER_DATA.nome if len(ERP_USER_DATA.nome) <= 18 else ERP_USER_DATA.nome[:16] + "...",
                            size=10,
                            weight=ft.FontWeight.W_500,
                            color=COR_PRIMARIA
                        ),
                    ], spacing=5),
                    ft.Text(
                        ERP_USER_DATA.cargo if ERP_USER_DATA.cargo else "Usuário",
                        size=9,
                        color=ft.Colors.GREY_600,
                        italic=True
                    ),
                    ft.TextButton(
                        "Voltar ao ERP",
                        icon=ft.Icons.ARROW_BACK,
                        on_click=self._voltar_ao_erp,
                        style=ft.ButtonStyle(
                            color=COR_PRIMARIA,
                            padding=ft.padding.all(0),
                        ),
                    ),
                ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.START),
                padding=8, bgcolor="#fff3e0", width=180,
                border=ft.border.only(bottom=ft.BorderSide(1, "#e67e22")),
            )

        return ft.Column([
            ft.Container(
                content=ft.Column([
                    ft.Image(src=logo_path, width=140, height=60, fit=ft.ImageFit.CONTAIN) if os.path.exists(logo_path) else ft.Icon(ft.Icons.BUSINESS, size=40, color=COR_PRIMARIA),
                    ft.Text("Gestão de RH", size=11, color=COR_PRIMARIA),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=5),
                padding=15, bgcolor="#e8f4fc", width=180,
            ),
            # Seção do ERP (se veio do ERP) - substitui a seção de usuário local
            erp_section if erp_section else ft.Container(
                # Seção do usuário logado (apenas se NÃO veio do ERP)
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.ACCOUNT_CIRCLE, size=24, color=COR_PRIMARIA),
                        ft.Column([
                            ft.Text(nome_usuario[:16] + ('...' if len(nome_usuario) > 16 else ''),
                                   size=10, weight=ft.FontWeight.BOLD, color=COR_PRIMARIA),
                            ft.Text(nivel_texto, size=9, color=ft.Colors.GREY_600),
                        ], spacing=0, expand=True),
                    ], spacing=6),
                    ft.Row([
                        ft.Container(
                            content=ft.Row([
                                ft.Icon(ft.Icons.PERSON, size=14, color=COR_SECUNDARIA),
                                ft.Text("Conta", size=10, color=COR_SECUNDARIA),
                            ], spacing=2),
                            on_click=lambda e: self._mostrar_minha_conta(),
                            ink=True,
                            padding=ft.padding.symmetric(horizontal=6, vertical=4),
                            border_radius=4,
                        ),
                        ft.Container(
                            content=ft.Row([
                                ft.Icon(ft.Icons.LOGOUT, size=14, color=COR_ERRO),
                                ft.Text("Sair", size=10, color=COR_ERRO),
                            ], spacing=2),
                            on_click=lambda e: self._fazer_logout(),
                            ink=True,
                            padding=ft.padding.symmetric(horizontal=6, vertical=4),
                            border_radius=4,
                        ),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, spacing=0),
                ], spacing=5),
                padding=8, bgcolor="#f5f5f5", width=180,
                border=ft.border.only(bottom=ft.BorderSide(1, COR_SECUNDARIA)),
            ),
            ft.Container(
                content=ft.Column(itens_menu, spacing=0, scroll=ft.ScrollMode.AUTO),
                padding=ft.padding.only(top=10), bgcolor="white", width=180, expand=True,
            ),
            ft.Container(
                content=ft.Column([
                    ft.Divider(color=COR_SECUNDARIA),
                    ft.Text("v1.05", size=10, color=ft.Colors.GREY),
                    ft.Text(f"© {datetime.now().year} RENOVO", size=10, color=ft.Colors.GREY),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=5),
                padding=10, bgcolor="white", width=180,
            ),
        ], spacing=0, width=180)
    
    def _item_menu(self, texto: str, icone, view: str):
        # Verificar se este item é a view atual
        is_ativo = hasattr(self, 'view_atual') and self.view_atual == view

        # Cor de fundo e texto baseado no estado
        if is_ativo:
            bgcolor = "#e67e22"  # Laranja para item ativo
            icon_color = "white"
            text_color = "white"
            font_weight = ft.FontWeight.BOLD
        else:
            bgcolor = None
            icon_color = COR_PRIMARIA
            text_color = None
            font_weight = None

        return ft.Container(
            content=ft.Row([
                ft.Icon(icone, size=20, color=icon_color),
                ft.Text(texto, size=13, color=text_color, weight=font_weight),
            ], spacing=10),
            padding=ft.padding.symmetric(horizontal=15, vertical=12),
            bgcolor=bgcolor,
            border_radius=ft.border_radius.only(top_left=8, bottom_left=8) if is_ativo else None,
            on_click=lambda e, v=view: self.navegar(v),
            ink=True,
        )
    
    def navegar(self, view: str):
        if view == "exportar":
            self._exportar_excel()
            return
        elif view == "backup":
            self._fazer_backup()
            return
        elif view == "sobre":
            self._mostrar_sobre()
            return
        elif view == "log":
            self._mostrar_log()
            return

        # Se estiver saindo do cadastro de novo colaborador, salvar dados temporários
        if self.view_atual == "cadastro" and self._formulario_atual:
            # Coletar e salvar dados do formulário atual
            try:
                dados_temp = self._formulario_atual._coletar_dados_temp()
                self._salvar_dados_temp(
                    dados_temp,
                    self._formulario_atual.dependentes_lista,
                    self._formulario_atual.foto_path
                )
            except (AttributeError, TypeError, ValueError) as e:
                # Log do erro mas continua execução para não bloquear navegação
                db.registrar_log("sistema", "aviso", f"Erro ao salvar dados temporários: {str(e)}")
            finally:
                self._formulario_atual = None  # Limpar referência

        self.view_atual = view
        self.atualizar_view()
    
    def atualizar_view(self, **kwargs):
        if self.view_atual == "lista":
            self.container_principal.content = self._view_lista_colaboradores()
        elif self.view_atual == "cadastro":
            editar_id = kwargs.get('editar_id')
            form = FormularioCadastro(
                self.page,
                on_salvar=self._on_formulario_salvar,
                on_cancelar=self._on_formulario_cancelar,
                on_salvar_temp=self._salvar_dados_temp,
                on_limpar=self._limpar_dados_temp,
                colaborador_id=editar_id,
                dados_temp=self.dados_formulario_temp if not editar_id else None,
                dependentes_temp=self.dependentes_temp if not editar_id else None,
                foto_path_temp=self.foto_path_temp if not editar_id else None,
            )
            # Guardar referência ao formulário (apenas para novo colaborador)
            self._formulario_atual = form if not editar_id else None
            self.container_principal.content = form.build()
        elif self.view_atual == "ficha":
            colaborador_id = kwargs.get('colaborador_id')
            ficha = FichaColaborador(self.page, colaborador_id, on_voltar=self._voltar_da_ficha)
            self.container_principal.content = ficha.build()
        elif self.view_atual == "contratos":
            self.container_principal.content = self._view_contratos()
        elif self.view_atual == "ferias":
            self.container_principal.content = self._view_ferias()
        elif self.view_atual == "aniversariantes":
            self.container_principal.content = self._view_aniversariantes()
        elif self.view_atual == "banco_talentos":
            self.container_principal.content = self._view_banco_talentos()
        elif self.view_atual == "blocklist":
            self.container_principal.content = self._view_blocklist()
        elif self.view_atual == "empresas":
            self.container_principal.content = self._view_empresas()
        elif self.view_atual == "dashboard":
            self.container_principal.content = self._view_dashboard()

        # Atualizar menu lateral para destacar item ativo (se já existir)
        if hasattr(self, 'layout_principal') and self.layout_principal:
            self.menu_lateral = self._criar_menu_lateral()
            self.layout_principal.controls[0] = self.menu_lateral

        self.page.update()
    
    def _voltar_da_ficha(self, **kwargs):
        try:
            if kwargs.get('editar_id'):
                self.view_atual = "cadastro"
                self.atualizar_view(editar_id=kwargs['editar_id'])
            elif kwargs.get('visualizar_id'):
                self.view_atual = "ficha"
                self.atualizar_view(colaborador_id=kwargs['visualizar_id'])
            else:
                # Voltar para a view anterior (se veio de contratos, volta para contratos)
                if hasattr(self, 'view_anterior') and self.view_anterior:
                    self.view_atual = self.view_anterior
                    self.view_anterior = None
                    self.atualizar_view()
                else:
                    self.navegar("lista")
        except (KeyError, AttributeError, TypeError) as e:
            # Log do erro e navegar para lista como fallback seguro
            db.registrar_log("sistema", "aviso", f"Erro ao voltar da ficha: {str(e)}")
            self.navegar("lista")

    def _abrir_ficha_colaborador(self, colaborador_id: int):
        """Abre a ficha de um colaborador."""
        # Guardar view atual para voltar depois
        self.view_anterior = self.view_atual
        self.view_atual = "ficha"
        self.atualizar_view(colaborador_id=colaborador_id)

    def _salvar_dados_temp(self, dados: dict, dependentes: list, foto_path: str):
        """Salva os dados do formulário temporariamente (em memória e em arquivo)."""
        self.dados_formulario_temp = dados
        self.dependentes_temp = dependentes
        self.foto_path_temp = foto_path
        # Persistir em arquivo para não perder ao fechar o programa
        self._salvar_dados_temp_arquivo()

    def _limpar_dados_temp(self):
        """Limpa os dados temporários do formulário (memória e arquivo)."""
        self.dados_formulario_temp = {}
        self.dependentes_temp = []
        self.foto_path_temp = None
        # Remover arquivo de persistência
        self._remover_arquivo_dados_temp()

    def _salvar_dados_temp_arquivo(self):
        """Persiste os dados temporários em arquivo JSON."""
        try:
            dados_persistir = {
                'dados_formulario': self.dados_formulario_temp,
                'dependentes': self.dependentes_temp,
                'foto_path': self.foto_path_temp
            }
            with open(self.ARQUIVO_DADOS_TEMP, 'w', encoding='utf-8') as f:
                json.dump(dados_persistir, f, ensure_ascii=False, indent=2)
        except Exception as e:
            db.registrar_log("sistema", "aviso", f"Erro ao persistir dados temporários: {str(e)}")

    def _carregar_dados_temp_arquivo(self):
        """Carrega os dados temporários do arquivo JSON (se existir)."""
        try:
            if os.path.exists(self.ARQUIVO_DADOS_TEMP):
                with open(self.ARQUIVO_DADOS_TEMP, 'r', encoding='utf-8') as f:
                    dados = json.load(f)
                    self.dados_formulario_temp = dados.get('dados_formulario', {})
                    self.dependentes_temp = dados.get('dependentes', [])
                    self.foto_path_temp = dados.get('foto_path')
        except Exception as e:
            db.registrar_log("sistema", "aviso", f"Erro ao carregar dados temporários: {str(e)}")
            # Se houver erro, inicializar com valores vazios
            self.dados_formulario_temp = {}
            self.dependentes_temp = []
            self.foto_path_temp = None

    def _remover_arquivo_dados_temp(self):
        """Remove o arquivo de dados temporários."""
        try:
            if os.path.exists(self.ARQUIVO_DADOS_TEMP):
                os.remove(self.ARQUIVO_DADOS_TEMP)
        except Exception as e:
            db.registrar_log("sistema", "aviso", f"Erro ao remover arquivo temporário: {str(e)}")

    def _on_formulario_salvar(self):
        """Callback quando o formulário é salvo com sucesso."""
        # Limpar dados temporários após salvar
        self._limpar_dados_temp()
        self.navegar("lista")

    def _on_formulario_cancelar(self):
        """Callback quando o usuário sai do formulário sem salvar."""
        # Os dados já foram salvos temporariamente pelo formulário
        self.navegar("lista")

    def _view_lista_colaboradores(self):
        self.empresa_selecionada = None
        self.localizacao_selecionada = None
        self.visualizando_inativos = False

        # Configurações de paginação
        self.pagina_atual = 1
        self.itens_por_pagina = 50
        self.total_colaboradores = 0

        # Carregar lista de empresas para o filtro
        empresas = db.listar_empresas(apenas_ativas=True)
        opcoes_empresas = [ft.dropdown.Option(key="", text="Todas as Empresas")]
        for emp in empresas:
            opcoes_empresas.append(ft.dropdown.Option(key=str(emp['id']), text=emp['razao_social']))

        self.dropdown_empresa = ft.Dropdown(
            label="Filtrar por Empresa",
            width=250,
            options=opcoes_empresas,
            value="",
            on_change=self._filtrar_por_empresa,
            border_color=COR_SECUNDARIA,
        )

        # Carregar lista de localizações para o filtro
        locais = db.contar_colaboradores_por_local()
        opcoes_locais = [ft.dropdown.Option(key="", text="Todas as Localizações")]
        for loc in locais:
            loc_texto = f"{loc['local_nome']} ({loc['qtd_colaboradores']})"
            opcoes_locais.append(ft.dropdown.Option(key=loc['local_nome'], text=loc_texto))

        self.dropdown_localizacao = ft.Dropdown(
            label="Filtrar por Localização",
            width=250,
            options=opcoes_locais,
            value="",
            on_change=self._filtrar_por_localizacao,
            border_color=COR_SECUNDARIA,
        )

        self.campo_pesquisa = ft.TextField(
            label="Pesquisar por nome ou CPF",
            prefix_icon=ft.Icons.SEARCH,
            width=300,
            on_change=self._filtrar_colaboradores,
            border_color=COR_SECUNDARIA,
        )

        # Botão para alternar entre ativos e inativos
        self.btn_toggle_inativos = ft.ElevatedButton(
            "Ver Inativos",
            icon=ft.Icons.PERSON_OFF,
            on_click=self._toggle_visualizar_inativos,
            bgcolor=COR_ALERTA,
            color="white",
        )

        self.lista_colaboradores = ft.Column(spacing=5, scroll=ft.ScrollMode.AUTO)
        self.titulo_lista = ft.Text("Colaboradores", size=24, weight=ft.FontWeight.BOLD)

        # Controles de paginação
        self.texto_paginacao = ft.Text("", size=13, color=COR_PRIMARIA)
        self.btn_pagina_anterior = ft.IconButton(
            icon=ft.Icons.CHEVRON_LEFT,
            icon_color=COR_PRIMARIA,
            tooltip="Página anterior",
            on_click=self._pagina_anterior,
            disabled=True,
        )
        self.btn_pagina_proxima = ft.IconButton(
            icon=ft.Icons.CHEVRON_RIGHT,
            icon_color=COR_PRIMARIA,
            tooltip="Próxima página",
            on_click=self._pagina_proxima,
            disabled=True,
        )

        self._carregar_colaboradores()

        return ft.Column([
            criar_alertas_widget(
                on_click_contrato=lambda e: self.navegar("contratos"),
                on_click_ferias=lambda e: self.navegar("ferias"),
            ),
            ft.Container(
                content=ft.Row([
                    ft.Row([
                        ft.Icon(ft.Icons.PEOPLE, size=30, color=COR_PRIMARIA),
                        self.titulo_lista,
                    ], spacing=10),
                    ft.Row([
                        self.btn_toggle_inativos,
                        self.dropdown_empresa,
                        self.dropdown_localizacao,
                        self.campo_pesquisa,
                        ft.ElevatedButton("Novo Colaborador", icon=ft.Icons.ADD,
                                         on_click=lambda e: self.navegar("cadastro"),
                                         bgcolor=COR_SUCESSO, color="white"),
                        ft.ElevatedButton("Relatório Localizações", icon=ft.Icons.LOCATION_ON,
                                         on_click=lambda e: self._exportar_localizacoes(),
                                         bgcolor="#e67e22", color="white"),
                    ], spacing=10, wrap=True),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                padding=15, bgcolor="white", border_radius=8,
            ),
            ft.Container(
                content=ft.Row([
                    ft.Container(content=ft.Text("", weight=ft.FontWeight.BOLD, color="white"), width=50),
                    ft.Container(content=ft.Text("Nome", weight=ft.FontWeight.BOLD, color="white"), expand=2),
                    ft.Container(content=ft.Text("CPF", weight=ft.FontWeight.BOLD, color="white"), expand=1),
                    ft.Container(content=ft.Text("Localização", weight=ft.FontWeight.BOLD, color="white"), expand=2),
                    ft.Container(content=ft.Text("Contrato", weight=ft.FontWeight.BOLD, color="white"), width=100, alignment=ft.alignment.center),
                    ft.Container(content=ft.Text("Docs", weight=ft.FontWeight.BOLD, color="white"), width=70, alignment=ft.alignment.center),
                    ft.Container(content=ft.Text("Empresa", weight=ft.FontWeight.BOLD, color="white"), expand=1),
                    ft.Container(content=ft.Text("Ações", weight=ft.FontWeight.BOLD, color="white"), width=120),
                ], spacing=15),
                padding=ft.padding.symmetric(horizontal=20, vertical=12),
                bgcolor=COR_PRIMARIA,
                border_radius=ft.border_radius.only(top_left=8, top_right=8),
            ),
            ft.Container(
                content=self.lista_colaboradores,
                border=ft.border.all(1, COR_SECUNDARIA),
                border_radius=ft.border_radius.only(bottom_left=8, bottom_right=8),
                expand=True,
            ),
            # Controles de paginação
            ft.Container(
                content=ft.Row([
                    self.texto_paginacao,
                    ft.Row([
                        self.btn_pagina_anterior,
                        self.btn_pagina_proxima,
                    ], spacing=0),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                padding=ft.padding.symmetric(horizontal=15, vertical=10),
                bgcolor="white",
                border_radius=8,
            ),
        ], spacing=10, expand=True)

    def _pagina_anterior(self, e):
        """Vai para a página anterior."""
        if self.pagina_atual > 1:
            self.pagina_atual -= 1
            filtro = self.campo_pesquisa.value if hasattr(self, 'campo_pesquisa') and self.campo_pesquisa.value else None
            self._carregar_colaboradores(filtro)
            self.page.update()

    def _pagina_proxima(self, e):
        """Vai para a próxima página."""
        total_paginas = (self.total_colaboradores + self.itens_por_pagina - 1) // self.itens_por_pagina
        if self.pagina_atual < total_paginas:
            self.pagina_atual += 1
            filtro = self.campo_pesquisa.value if hasattr(self, 'campo_pesquisa') and self.campo_pesquisa.value else None
            self._carregar_colaboradores(filtro)
            self.page.update()
    
    def _carregar_colaboradores(self, filtro: str = None):
        empresa_id = None
        if hasattr(self, 'empresa_selecionada') and self.empresa_selecionada:
            empresa_id = int(self.empresa_selecionada)

        localizacao = None
        if hasattr(self, 'localizacao_selecionada') and self.localizacao_selecionada:
            localizacao = self.localizacao_selecionada

        # Determinar status baseado no modo de visualização
        status_filtro = 'INATIVO' if hasattr(self, 'visualizando_inativos') and self.visualizando_inativos else 'ATIVO'

        # Calcular offset para paginação
        offset = (self.pagina_atual - 1) * self.itens_por_pagina

        # Obter total de colaboradores e lista paginada
        self.total_colaboradores = db.contar_colaboradores(filtro=filtro, status=status_filtro, empresa_id=empresa_id, localizacao=localizacao)
        colaboradores = db.listar_colaboradores(
            filtro=filtro,
            status=status_filtro,
            empresa_id=empresa_id,
            localizacao=localizacao,
            limite=self.itens_por_pagina,
            offset=offset
        )

        # Atualizar controles de paginação
        total_paginas = max(1, (self.total_colaboradores + self.itens_por_pagina - 1) // self.itens_por_pagina)
        inicio = offset + 1 if self.total_colaboradores > 0 else 0
        fim = min(offset + self.itens_por_pagina, self.total_colaboradores)

        self.texto_paginacao.value = f"Exibindo {inicio}-{fim} de {self.total_colaboradores} colaboradores (Página {self.pagina_atual} de {total_paginas})"
        self.btn_pagina_anterior.disabled = self.pagina_atual <= 1
        self.btn_pagina_proxima.disabled = self.pagina_atual >= total_paginas

        self.lista_colaboradores.controls.clear()

        if not colaboradores:
            msg = "Nenhum colaborador inativo encontrado" if status_filtro == 'INATIVO' else "Nenhum colaborador encontrado"
            self.lista_colaboradores.controls.append(
                ft.Container(
                    content=ft.Text(msg, italic=True, color=ft.Colors.GREY),
                    padding=20, alignment=ft.alignment.center,
                )
            )
        else:
            for colab in colaboradores:
                empresa_nome = colab.get('empresa_nome', '') or ''
                foto_path = colab.get('foto_path', '')

                # Obter localização atual do colaborador
                localizacao = db.obter_localizacao_atual(colab['id'])
                if localizacao:
                    loc_texto = f"{localizacao.get('local_nome', '')}"
                    if localizacao.get('cidade') or localizacao.get('uf'):
                        loc_texto += f" - {localizacao.get('cidade', '')}/{localizacao.get('uf', '')}"
                else:
                    loc_texto = "-"

                # Criar widget de foto
                cor_avatar = COR_SECUNDARIA if status_filtro == 'ATIVO' else COR_ERRO
                foto_widget = self._criar_avatar_foto(foto_path, cor_avatar)

                # Determinar tipo de contrato e cor do badge
                tipo_contrato = colab.get('tipo_contrato', '') or ''
                if tipo_contrato == 'Contrato de Experiência':
                    contrato_texto = "Experiência"
                    contrato_cor = COR_ALERTA
                elif tipo_contrato == 'CLT':
                    contrato_texto = "CLT"
                    contrato_cor = COR_SUCESSO
                elif tipo_contrato:
                    contrato_texto = tipo_contrato[:12]  # Truncar se muito longo
                    contrato_cor = COR_SECUNDARIA
                else:
                    contrato_texto = "-"
                    contrato_cor = COR_CINZA_CLARO

                # Widget do badge de contrato
                contrato_widget = ft.Container(
                    content=ft.Text(contrato_texto, color="white", size=11),
                    bgcolor=contrato_cor,
                    padding=ft.padding.symmetric(horizontal=8, vertical=3),
                    border_radius=4,
                ) if contrato_texto != "-" else ft.Text("-", size=14, color=ft.Colors.GREY)

                # Widget do badge de documentos
                status_docs = db.obter_status_documentos_colaborador(colab['id'])
                docs_completos = status_docs.get('completos', 0)
                docs_total = status_docs.get('total_obrigatorios', 0)

                if docs_total > 0:
                    if docs_completos == docs_total:
                        docs_cor = COR_SUCESSO  # Verde - completo
                    elif docs_completos >= docs_total * 0.5:
                        docs_cor = COR_ALERTA  # Amarelo - parcial
                    else:
                        docs_cor = COR_ERRO  # Vermelho - poucos docs

                    docs_widget = ft.Container(
                        content=ft.Text(f"{docs_completos}/{docs_total}", color="white", size=11),
                        bgcolor=docs_cor,
                        padding=ft.padding.symmetric(horizontal=8, vertical=3),
                        border_radius=4,
                        tooltip=f"{docs_completos} de {docs_total} documentos anexados",
                    )
                else:
                    docs_widget = ft.Text("-", size=14, color=ft.Colors.GREY)

                # Botões de ação - diferentes para ativos e inativos
                if status_filtro == 'INATIVO':
                    botoes_acao = ft.Row([
                        ft.IconButton(icon=ft.Icons.PERSON_ADD, icon_color=COR_SUCESSO,
                                     tooltip="Reativar", on_click=lambda e, c=colab: self._reativar_colaborador(c)),
                    ], spacing=0)
                else:
                    botoes_acao = ft.Row([
                        ft.IconButton(icon=ft.Icons.PICTURE_AS_PDF, icon_color=COR_ERRO,
                                     tooltip="PDF", on_click=lambda e, c=colab: self._gerar_pdf(c)),
                    ], spacing=0)

                self.lista_colaboradores.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Container(content=foto_widget, width=50, alignment=ft.alignment.center),
                            ft.Container(content=ft.Text(colab.get('nome_completo', ''), size=14), expand=2),
                            ft.Container(content=ft.Text(formatar_cpf(colab.get('cpf', '')), size=14), expand=1),
                            ft.Container(content=ft.Text(loc_texto, size=14, color=ft.Colors.GREY_700 if loc_texto == "-" else None), expand=2),
                            ft.Container(content=contrato_widget, width=100, alignment=ft.alignment.center),
                            ft.Container(content=docs_widget, width=70, alignment=ft.alignment.center),
                            ft.Container(content=ft.Text(empresa_nome, size=14), expand=1),
                            ft.Container(content=botoes_acao, width=120),
                        ], spacing=15),
                        padding=ft.padding.symmetric(horizontal=20, vertical=8),
                        bgcolor=ft.Colors.WHITE if status_filtro == 'ATIVO' else "#FFF5F5",
                        border=ft.border.only(bottom=ft.BorderSide(1, COR_CINZA_CLARO)),
                        on_click=lambda e, c=colab: self._abrir_ficha(c),
                        ink=True,
                    )
                )

    def _toggle_visualizar_inativos(self, e):
        """Alterna entre visualização de colaboradores ativos e inativos."""
        self.visualizando_inativos = not self.visualizando_inativos
        # Resetar para primeira página ao alternar
        self.pagina_atual = 1

        if self.visualizando_inativos:
            self.btn_toggle_inativos.text = "Ver Ativos"
            self.btn_toggle_inativos.icon = ft.Icons.PERSON
            self.btn_toggle_inativos.bgcolor = COR_SUCESSO
            self.titulo_lista.value = "Colaboradores Inativos"
        else:
            self.btn_toggle_inativos.text = "Ver Inativos"
            self.btn_toggle_inativos.icon = ft.Icons.PERSON_OFF
            self.btn_toggle_inativos.bgcolor = COR_ALERTA
            self.titulo_lista.value = "Colaboradores"

        filtro_texto = self.campo_pesquisa.value if hasattr(self, 'campo_pesquisa') else None
        self._carregar_colaboradores(filtro_texto)
        self.page.update()

    def _reativar_colaborador(self, colaborador):
        """Abre diálogo para reativar um colaborador inativo."""

        def confirmar(ev):
            try:
                # Atualizar status para ATIVO
                db.atualizar_colaborador(colaborador['id'], {'status': 'ATIVO', 'motivo_inativacao': None, 'submotivo_inativacao': None})

                # Registrar no histórico
                db.registrar_alteracao(colaborador['id'], 'status', 'INATIVO', 'ATIVO')

                dialog.open = False
                self.page.snack_bar = ft.SnackBar(
                    content=ft.Text(f"Colaborador {colaborador.get('nome_completo', '')} reativado com sucesso!"),
                    bgcolor=COR_SUCESSO
                )
                self.page.snack_bar.open = True

                # Recarregar lista
                filtro_texto = self.campo_pesquisa.value if hasattr(self, 'campo_pesquisa') else None
                self._carregar_colaboradores(filtro_texto)
                self.page.update()
            except (ValueError, KeyError, TypeError) as ex:
                db.registrar_log("sistema", "erro", f"Erro ao reativar colaborador: {str(ex)}")
                self.page.snack_bar = ft.SnackBar(content=ft.Text(f"Erro ao reativar: {str(ex)}"), bgcolor=COR_ERRO)
                self.page.snack_bar.open = True
                self.page.update()

        def cancelar(ev):
            dialog.open = False
            self.page.update()

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.PERSON_ADD, color=COR_SUCESSO, size=28),
                ft.Text("Reativar Colaborador"),
            ]),
            content=ft.Column([
                ft.Text(f"Deseja reativar o colaborador:", size=14),
                ft.Text(colaborador.get('nome_completo', ''), weight=ft.FontWeight.BOLD, size=16),
                ft.Text(f"CPF: {formatar_cpf(colaborador.get('cpf', ''))}", size=13, color=ft.Colors.GREY_700),
            ], spacing=5),
            actions=[
                ft.TextButton("Cancelar", on_click=cancelar),
                ft.ElevatedButton("Reativar", on_click=confirmar, bgcolor=COR_SUCESSO, color="white"),
            ],
        )

        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def _filtrar_colaboradores(self, e):
        # Resetar para primeira página ao filtrar
        self.pagina_atual = 1
        self._carregar_colaboradores(e.control.value)
        self.page.update()

    def _filtrar_por_empresa(self, e):
        self.empresa_selecionada = e.control.value if e.control.value else None
        # Resetar para primeira página ao mudar empresa
        self.pagina_atual = 1
        filtro_texto = self.campo_pesquisa.value if hasattr(self, 'campo_pesquisa') else None
        self._carregar_colaboradores(filtro_texto)
        self.page.update()

    def _filtrar_por_localizacao(self, e):
        self.localizacao_selecionada = e.control.value if e.control.value else None
        # Resetar para primeira página ao mudar localização
        self.pagina_atual = 1
        filtro_texto = self.campo_pesquisa.value if hasattr(self, 'campo_pesquisa') else None
        self._carregar_colaboradores(filtro_texto)
        self.page.update()

    def _abrir_ficha(self, colaborador):
        self.view_atual = "ficha"
        self.atualizar_view(colaborador_id=colaborador['id'])
    
    def _gerar_pdf(self, colaborador):
        try:
            colab = db.obter_colaborador(colaborador['id'])
            deps = db.listar_dependentes(colaborador['id'])
            empresa = db.obter_empresa(colab.get('empresa_id')) if colab.get('empresa_id') else None
            output = gerar_ficha_registro_pdf(colab, deps, empresa)

            self.page.snack_bar = ft.SnackBar(content=ft.Text(f"PDF gerado: {output}"), bgcolor=COR_SUCESSO)
            self.page.snack_bar.open = True
            self.page.update()

            # Abrir o PDF automaticamente
            os.startfile(output)
        except (IOError, OSError, ValueError) as ex:
            self.page.snack_bar = ft.SnackBar(content=ft.Text(f"Erro ao gerar PDF: {str(ex)}"), bgcolor=COR_ERRO)
            self.page.snack_bar.open = True
            self.page.update()

    def _exportar_localizacoes(self):
        """Exporta relatório de colaboradores por localização para Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from datetime import datetime as dt
            import os

            # Obter dados agrupados por localização
            locais = db.contar_colaboradores_por_local()

            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resumo por Localização"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Título
            ws['A1'] = f"Relatório de Colaboradores por Localização - {dt.now().strftime('%d/%m/%Y')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')

            # Cabeçalhos - Resumo
            headers = ['Local/Empresa', 'Cidade', 'UF', 'Qtd. Colaboradores']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Dados do resumo
            row = 4
            total_geral = 0
            for loc in locais:
                ws.cell(row=row, column=1, value=loc.get('local_nome', '')).border = thin_border
                ws.cell(row=row, column=2, value=loc.get('cidade', '') or '').border = thin_border
                ws.cell(row=row, column=3, value=loc.get('uf', '') or '').border = thin_border
                ws.cell(row=row, column=4, value=loc.get('qtd_colaboradores', 0)).border = thin_border
                total_geral += loc.get('qtd_colaboradores', 0)
                row += 1

            # Total
            ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=row, column=4, value=total_geral).font = Font(bold=True)

            # Ajustar larguras
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 18

            # Segunda aba - Lista detalhada
            ws2 = wb.create_sheet("Detalhado por Local")

            # Para cada localização, listar os colaboradores
            row2 = 1
            for loc in locais:
                # Título do local
                ws2.cell(row=row2, column=1, value=f"{loc.get('local_nome', '')} - {loc.get('cidade', '') or ''}/{loc.get('uf', '') or ''}")
                ws2.cell(row=row2, column=1).font = Font(bold=True, size=12)
                ws2.merge_cells(start_row=row2, start_column=1, end_row=row2, end_column=5)
                row2 += 1

                # Cabeçalhos
                headers_det = ['Nome', 'CPF', 'Função', 'Empresa', 'Desde']
                for col, header in enumerate(headers_det, 1):
                    cell = ws2.cell(row=row2, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                row2 += 1

                # Buscar colaboradores desse local
                colaboradores = db.listar_colaboradores_por_localizacao(local_nome=loc.get('local_nome'))
                for colab in colaboradores:
                    ws2.cell(row=row2, column=1, value=colab.get('nome_completo', '')).border = thin_border
                    ws2.cell(row=row2, column=2, value=colab.get('cpf', '')).border = thin_border
                    ws2.cell(row=row2, column=3, value=colab.get('funcao', '') or '').border = thin_border
                    ws2.cell(row=row2, column=4, value=colab.get('empresa_nome', '') or '').border = thin_border
                    # Formatar data
                    data_loc = colab.get('loc_data_inicio', '')
                    if data_loc:
                        try:
                            data_formatada = dt.strptime(data_loc, '%Y-%m-%d').strftime('%d/%m/%Y')
                        except (ValueError, TypeError):
                            data_formatada = data_loc
                    else:
                        data_formatada = ''
                    ws2.cell(row=row2, column=5, value=data_formatada).border = thin_border
                    row2 += 1

                row2 += 1  # Linha em branco entre locais

            # Ajustar larguras da aba detalhada
            ws2.column_dimensions['A'].width = 40
            ws2.column_dimensions['B'].width = 15
            ws2.column_dimensions['C'].width = 25
            ws2.column_dimensions['D'].width = 30
            ws2.column_dimensions['E'].width = 12

            # Salvar arquivo
            output_dir = os.path.join(os.path.expanduser('~'), 'Documents', 'RH_Sistema')
            os.makedirs(output_dir, exist_ok=True)
            output_file = os.path.join(output_dir, f"relatorio_localizacoes_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb.save(output_file)

            self.page.snack_bar = ft.SnackBar(
                content=ft.Text(f"Relatório exportado: {output_file}"),
                bgcolor=COR_SUCESSO
            )
            self.page.snack_bar.open = True
            self.page.update()

            # Abrir o arquivo automaticamente
            os.startfile(output_file)

        except (IOError, OSError, ValueError) as ex:
            self.page.snack_bar = ft.SnackBar(
                content=ft.Text(f"Erro ao exportar: {str(ex)}"),
                bgcolor=COR_ERRO
            )
            self.page.snack_bar.open = True
            self.page.update()

    def _view_contratos(self):
        contratos = db.listar_todos_contratos_experiencia()
        lista = []

        for c in contratos:
            dias = c.get('dias_restantes', 0)
            data_venc = c.get('proxima_data_vencimento', '')

            # Definir cor baseada na urgência
            if dias <= 0:
                cor = COR_ERRO
                status_texto = "VENCIDO"
            elif dias <= 5:
                cor = COR_ERRO
                status_texto = f"{dias} dias"
            elif dias <= 15:
                cor = COR_ALERTA
                status_texto = f"{dias} dias"
            else:
                cor = COR_SUCESSO
                status_texto = f"{dias} dias"

            # Texto do período - baseado se tem prorrogação ou não
            prorrogacao = c.get('prorrogacao')
            prazo_inicial = c.get('prazo_inicial', 0)

            if prorrogacao:
                # Tem prorrogação = está no 2º período
                periodo_texto = "2º Período"
                periodo_info = f"Inicial: {prazo_inicial} dias + Prorrogação: {prorrogacao} dias = {prazo_inicial + prorrogacao} dias"
            else:
                # Sem prorrogação = 1º período
                periodo_texto = "1º Período"
                periodo_info = f"Prazo: {prazo_inicial} dias (aguardando renovação)"

            lista.append(
                ft.Container(
                    content=ft.Row([
                        ft.Column([
                            ft.Text(c.get('nome_completo', ''), weight=ft.FontWeight.BOLD, size=15),
                            ft.Text(f"{c.get('funcao', '')} - {c.get('empresa_nome', '')}", size=12, color=ft.Colors.GREY),
                            ft.Row([
                                ft.Container(
                                    content=ft.Text(periodo_texto, color="white", size=11),
                                    bgcolor=COR_SECUNDARIA if not prorrogacao else COR_PRIMARIA,
                                    padding=ft.padding.symmetric(horizontal=8, vertical=2),
                                    border_radius=4,
                                ),
                                ft.Text(periodo_info, size=11, color=ft.Colors.GREY_700),
                            ], spacing=10),
                        ], spacing=4, expand=True),
                        ft.Column([
                            ft.Container(
                                content=ft.Text(status_texto, color="white", weight=ft.FontWeight.BOLD, size=14),
                                bgcolor=cor, padding=ft.padding.symmetric(horizontal=15, vertical=8), border_radius=6,
                            ),
                            ft.Text(f"Vence: {formatar_data_br(data_venc)}", size=12),
                            ft.Text(f"Início: {formatar_data_br(c.get('data_inicio', ''))}", size=11, color=ft.Colors.GREY),
                        ], horizontal_alignment=ft.CrossAxisAlignment.END, spacing=3),
                    ]),
                    padding=15, border=ft.border.all(1, cor), border_radius=8, bgcolor="white",
                    on_click=lambda e, colab_id=c.get('colaborador_id'): self._abrir_ficha_colaborador(colab_id),
                )
            )

        # Resumo
        total = len(contratos)
        vencidos = len([c for c in contratos if c.get('dias_restantes', 0) <= 0])
        urgentes = len([c for c in contratos if 0 < c.get('dias_restantes', 0) <= 5])
        proximos = len([c for c in contratos if 5 < c.get('dias_restantes', 0) <= 15])

        return ft.Column([
            ft.Container(
                content=ft.Row([
                    ft.Row([
                        ft.Icon(ft.Icons.SCHEDULE, size=30, color=COR_PRIMARIA),
                        ft.Text("Contratos de Experiência", size=24, weight=ft.FontWeight.BOLD),
                    ]),
                    ft.Row([
                        ft.Container(
                            content=ft.Text(f"{vencidos} vencidos", color="white", size=12),
                            bgcolor=COR_ERRO, padding=ft.padding.symmetric(horizontal=10, vertical=5), border_radius=4,
                            visible=vencidos > 0,
                        ),
                        ft.Container(
                            content=ft.Text(f"{urgentes} urgentes", color="white", size=12),
                            bgcolor=COR_ALERTA, padding=ft.padding.symmetric(horizontal=10, vertical=5), border_radius=4,
                            visible=urgentes > 0,
                        ),
                        ft.Container(
                            content=ft.Text(f"{total} total", color="white", size=12),
                            bgcolor=COR_SECUNDARIA, padding=ft.padding.symmetric(horizontal=10, vertical=5), border_radius=4,
                        ),
                        ft.ElevatedButton("Exportar", icon=ft.Icons.DOWNLOAD, on_click=lambda e: self._exportar_contratos(),
                                         bgcolor=COR_SECUNDARIA, color="white"),
                    ], spacing=10),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                padding=15, bgcolor="white", border_radius=8,
            ),
            ft.Container(
                content=ft.Text("Ordenado por dias restantes (mais urgentes primeiro)", size=12, italic=True, color=ft.Colors.GREY),
                padding=ft.padding.only(left=10, top=5),
            ),
            ft.Container(
                content=ft.Column(lista if lista else [
                    ft.Container(
                        content=ft.Column([
                            ft.Icon(ft.Icons.CHECK_CIRCLE, size=50, color=COR_SUCESSO),
                            ft.Text("Nenhum contrato de experiência vigente!", size=16, color=COR_SUCESSO),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                        padding=30,
                    )
                ], scroll=ft.ScrollMode.AUTO, spacing=10),
                expand=True, padding=10,
            ),
        ], spacing=10, expand=True)
    
    def _view_ferias(self):
        ferias = db.listar_ferias_vencendo(6)
        lista = []

        for f in ferias:
            limite = datetime.strptime(f.get('periodo_concessivo_limite'), '%Y-%m-%d')
            meses = (limite - datetime.now()).days // 30
            cor = COR_ERRO if meses <= 2 else COR_ALERTA
            colaborador_id = f.get('colaborador_id')

            # Calcular início do período concessivo (dia seguinte ao fim do aquisitivo)
            periodo_concessivo_inicio = datetime.strptime(f.get('periodo_aquisitivo_fim'), '%Y-%m-%d') + timedelta(days=1)

            # Montar texto das últimas férias
            ultimas_ferias_texto = "Sem férias anteriores"
            if f.get('ultimas_ferias_inicio') and f.get('ultimas_ferias_fim'):
                ultimas_ferias_texto = f"{formatar_data_br(f.get('ultimas_ferias_inicio'))} a {formatar_data_br(f.get('ultimas_ferias_fim'))} ({f.get('ultimas_ferias_dias')} dias)"

            lista.append(
                ft.Container(
                    content=ft.Row([
                        ft.Column([
                            ft.Text(f.get('nome_completo', ''), weight=ft.FontWeight.BOLD),
                            ft.Text(f"Últimas Férias: {ultimas_ferias_texto}", size=12, color=COR_SUCESSO),
                            ft.Text(f"Período Aquisitivo: {formatar_data_br(f.get('periodo_aquisitivo_inicio'))} a {formatar_data_br(f.get('periodo_aquisitivo_fim'))}", size=12),
                            ft.Text(f"Período Concessivo: {formatar_data_br(periodo_concessivo_inicio.strftime('%Y-%m-%d'))} a {formatar_data_br(f.get('periodo_concessivo_limite'))}", size=12, color=cor),
                        ], spacing=2, expand=True),
                        ft.Column([
                            ft.Container(
                                content=ft.Text(f"{'VENCIDO' if meses <= 0 else f'{meses} meses'}", color="white"),
                                bgcolor=cor, padding=ft.padding.symmetric(horizontal=10, vertical=5), border_radius=4,
                            ),
                            ft.Text(f"Limite: {formatar_data_br(f.get('periodo_concessivo_limite'))}", size=12),
                        ], horizontal_alignment=ft.CrossAxisAlignment.END, spacing=5),
                    ]),
                    padding=15, border=ft.border.all(1, cor), border_radius=8, bgcolor="white",
                    on_click=lambda e, cid=colaborador_id: self._abrir_ficha_colaborador(cid),
                    ink=True,
                )
            )

        # Contar colaboradores em férias hoje
        em_ferias = db.listar_colaboradores_em_ferias()
        qtd_em_ferias = len(em_ferias)

        return ft.Column([
            ft.Container(
                content=ft.Row([
                    ft.Row([
                        ft.Icon(ft.Icons.BEACH_ACCESS, size=30, color=COR_ALERTA),
                        ft.Text("Férias - Período Concessivo Vencendo", size=24, weight=ft.FontWeight.BOLD),
                    ], spacing=10),
                    ft.Row([
                        ft.ElevatedButton(
                            f"Em Férias Hoje ({qtd_em_ferias})",
                            icon=ft.Icons.FLIGHT_TAKEOFF,
                            on_click=lambda e: self._mostrar_em_ferias_hoje(),
                            bgcolor=COR_SUCESSO if qtd_em_ferias > 0 else ft.Colors.GREY,
                            color="white",
                        ),
                        ft.ElevatedButton("Exportar", icon=ft.Icons.DOWNLOAD, on_click=lambda e: self._exportar_ferias(),
                                         bgcolor=COR_SECUNDARIA, color="white"),
                    ], spacing=10),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                padding=15, bgcolor="white", border_radius=8,
            ),
            ft.Container(
                content=ft.Column(lista if lista else [ft.Text("Nenhuma férias vencendo!", color=COR_SUCESSO)],
                                 scroll=ft.ScrollMode.AUTO, spacing=10),
                expand=True, padding=10,
            ),
        ], spacing=10, expand=True)

    def _mostrar_em_ferias_hoje(self):
        """Mostra diálogo com os colaboradores que estão de férias hoje."""
        em_ferias = db.listar_colaboradores_em_ferias()

        def fechar(ev):
            dialog.open = False
            self.page.update()

        lista_widgets = []
        if not em_ferias:
            lista_widgets.append(
                ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.CHECK_CIRCLE, size=50, color=COR_SUCESSO),
                        ft.Text("Nenhum colaborador está de férias hoje.", size=14, color=ft.Colors.GREY),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                    padding=30,
                    alignment=ft.alignment.center,
                )
            )
        else:
            for colab in em_ferias:
                foto_path = colab.get('foto_path', '')

                # Calcular dias restantes para voltar
                try:
                    data_fim = datetime.strptime(colab.get('data_fim', ''), '%Y-%m-%d')
                    hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                    dias_restantes = (data_fim - hoje).days + 1  # +1 porque inclui o dia de hoje
                    if dias_restantes < 0:
                        dias_restantes = 0
                except (ValueError, TypeError, AttributeError):
                    dias_restantes = 0

                # Widget de foto
                foto_widget = self._criar_avatar_foto(foto_path, COR_SUCESSO)

                # Definir cor do badge de dias restantes
                if dias_restantes <= 3:
                    cor_dias_restantes = COR_ERRO
                elif dias_restantes <= 7:
                    cor_dias_restantes = COR_ALERTA
                else:
                    cor_dias_restantes = COR_SECUNDARIA

                lista_widgets.append(
                    ft.Container(
                        content=ft.Row([
                            foto_widget,
                            ft.Column([
                                ft.Text(colab.get('nome_completo', ''), weight=ft.FontWeight.BOLD, size=14),
                                ft.Text(f"{colab.get('funcao', '')} - {colab.get('empresa_nome', '')}", size=12, color=ft.Colors.GREY_700),
                            ], spacing=2, expand=True),
                            ft.Column([
                                ft.Row([
                                    ft.Container(
                                        content=ft.Text(f"{colab.get('dias', 0)} dias", color="white", size=10),
                                        bgcolor=COR_SUCESSO,
                                        padding=ft.padding.symmetric(horizontal=6, vertical=2),
                                        border_radius=4,
                                    ),
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(ft.Icons.ARROW_BACK, size=12, color="white"),
                                            ft.Text(f"{dias_restantes}d", color="white", size=10, weight=ft.FontWeight.BOLD),
                                        ], spacing=2),
                                        bgcolor=cor_dias_restantes,
                                        padding=ft.padding.symmetric(horizontal=6, vertical=2),
                                        border_radius=4,
                                        tooltip=f"Volta em {dias_restantes} dia(s)",
                                    ),
                                ], spacing=5),
                                ft.Text(
                                    f"{formatar_data_br(colab.get('data_inicio'))} a {formatar_data_br(colab.get('data_fim'))}",
                                    size=11, color=ft.Colors.GREY_600,
                                ),
                            ], horizontal_alignment=ft.CrossAxisAlignment.END, spacing=3),
                        ], spacing=15),
                        padding=12,
                        border=ft.border.all(1, COR_SUCESSO),
                        border_radius=8,
                        bgcolor="white",
                    )
                )

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.FLIGHT_TAKEOFF, color=COR_SUCESSO, size=28),
                ft.Text(f"Colaboradores em Férias Hoje ({len(em_ferias)})"),
            ]),
            content=ft.Container(
                content=ft.Column(lista_widgets, scroll=ft.ScrollMode.AUTO, spacing=10),
                width=520,
                height=350,
            ),
            actions=[ft.ElevatedButton("Fechar", on_click=fechar, bgcolor=COR_SECUNDARIA, color="white")],
        )

        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def _view_aniversariantes(self):
        mes = datetime.now().month
        meses = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
        aniversariantes = db.listar_aniversariantes_mes(mes)
        lista = []
        
        for a in aniversariantes:
            dia = ""
            try:
                dia = datetime.strptime(str(a['data_nascimento']), '%Y-%m-%d').day
            except (ValueError, TypeError, KeyError):
                pass
            
            lista.append(
                ft.ListTile(
                    leading=ft.CircleAvatar(content=ft.Text(str(dia)), bgcolor=COR_SECUNDARIA, color="white"),
                    title=ft.Text(a.get('nome_completo', ''), weight=ft.FontWeight.BOLD),
                    subtitle=ft.Text(f"{a.get('funcao', '')} - {a.get('empresa_nome', '')}"),
                )
            )
        
        return ft.Column([
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.CELEBRATION, size=30, color=COR_SUCESSO),
                    ft.Text(f"Aniversariantes de {meses[mes]}", size=24, weight=ft.FontWeight.BOLD),
                    ft.ElevatedButton("Exportar", icon=ft.Icons.DOWNLOAD, on_click=lambda e: self._exportar_aniversariantes(),
                                     bgcolor=COR_SECUNDARIA, color="white"),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                padding=15, bgcolor="white", border_radius=8,
            ),
            ft.Container(
                content=ft.Column(lista if lista else [ft.Text("Nenhum aniversariante!", italic=True)],
                                 scroll=ft.ScrollMode.AUTO),
                expand=True, padding=10, bgcolor="white", border_radius=8,
            ),
        ], spacing=10, expand=True)

    def _view_banco_talentos(self):
        """Exibe todos os colaboradores que já passaram pelo sistema."""
        stats = db.obter_estatisticas_banco_talentos()

        # Variáveis de filtro
        self.filtro_status_banco = None
        self.filtro_empresa_banco = None

        # Carregar lista de empresas para o filtro
        empresas = db.listar_empresas(apenas_ativas=False)
        opcoes_empresas = [ft.dropdown.Option(key="", text="Todas as Empresas")]
        for emp in empresas:
            opcoes_empresas.append(ft.dropdown.Option(key=str(emp['id']), text=emp['razao_social']))

        dropdown_empresa = ft.Dropdown(
            label="Empresa",
            options=opcoes_empresas,
            width=200,
            border_color=COR_SECUNDARIA,
            on_change=lambda e: self._filtrar_banco_talentos(e, 'empresa'),
        )

        dropdown_status = ft.Dropdown(
            label="Status",
            options=[
                ft.dropdown.Option(key="", text="Todos os Status"),
                ft.dropdown.Option(key="ativo", text="Ativos"),
                ft.dropdown.Option(key="inativo", text="Inativos"),
                ft.dropdown.Option(key="blocklist", text="Block-List"),
            ],
            width=150,
            border_color=COR_SECUNDARIA,
            on_change=lambda e: self._filtrar_banco_talentos(e, 'status'),
        )

        campo_pesquisa = ft.TextField(
            label="Pesquisar por nome ou CPF",
            prefix_icon=ft.Icons.SEARCH,
            width=250,
            on_change=lambda e: self._filtrar_banco_talentos(e, 'texto'),
            border_color=COR_SECUNDARIA,
        )

        # Guardar referências para uso nos filtros
        self.campo_pesquisa_banco = campo_pesquisa
        self.dropdown_empresa_banco = dropdown_empresa
        self.dropdown_status_banco = dropdown_status

        # Lista de colaboradores
        self.lista_banco_talentos = ft.Column(spacing=5, scroll=ft.ScrollMode.AUTO)
        self._carregar_banco_talentos()

        # Cards de estatísticas
        cards_stats = ft.Row([
            self._criar_card_stat("Total", stats['total'], ft.Icons.FOLDER_SHARED, COR_PRIMARIA),
            self._criar_card_stat("Ativos", stats['ativos'], ft.Icons.PERSON, COR_SUCESSO),
            self._criar_card_stat("Inativos", stats['inativos'], ft.Icons.PERSON_OFF, COR_ALERTA),
            self._criar_card_stat("Block-List", stats['blocklist'], ft.Icons.BLOCK, COR_ERRO),
        ], spacing=15)

        return ft.Column([
            ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Row([
                            ft.Icon(ft.Icons.FOLDER_SHARED, size=30, color=COR_PRIMARIA),
                            ft.Text("Banco de Talentos", size=24, weight=ft.FontWeight.BOLD),
                        ], spacing=10),
                    ]),
                    ft.Text(
                        "Registro completo de todos os colaboradores que já passaram pela empresa",
                        size=13,
                        color=ft.Colors.GREY_700,
                        italic=True,
                    ),
                    ft.Container(height=10),
                    cards_stats,
                ], spacing=5),
                padding=15,
                bgcolor="white",
                border_radius=8,
            ),
            ft.Container(
                content=ft.Row([
                    campo_pesquisa,
                    dropdown_status,
                    dropdown_empresa,
                ], spacing=15),
                padding=10,
            ),
            ft.Container(
                content=self.lista_banco_talentos,
                expand=True,
                padding=10,
                bgcolor="white",
                border_radius=8,
            ),
        ], spacing=10, expand=True)

    def _carregar_banco_talentos(self, filtro_texto: str = None):
        """Carrega a lista do banco de talentos."""
        empresa_id = None
        if hasattr(self, 'filtro_empresa_banco') and self.filtro_empresa_banco:
            empresa_id = int(self.filtro_empresa_banco)

        colaboradores = db.listar_banco_talentos(filtro=filtro_texto, empresa_id=empresa_id)

        # Aplicar filtro de status se definido
        if hasattr(self, 'filtro_status_banco') and self.filtro_status_banco:
            if self.filtro_status_banco == 'ativo':
                colaboradores = [c for c in colaboradores if c.get('status') == 'ATIVO']
            elif self.filtro_status_banco == 'inativo':
                colaboradores = [c for c in colaboradores if c.get('status') == 'INATIVO' and not c.get('na_blocklist')]
            elif self.filtro_status_banco == 'blocklist':
                colaboradores = [c for c in colaboradores if c.get('na_blocklist')]

        self.lista_banco_talentos.controls.clear()

        if not colaboradores:
            self.lista_banco_talentos.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.SEARCH_OFF, size=50, color=ft.Colors.GREY),
                        ft.Text("Nenhum colaborador encontrado", italic=True, color=ft.Colors.GREY),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                    padding=30,
                    alignment=ft.alignment.center,
                )
            )
        else:
            for colab in colaboradores:
                empresa_nome = colab.get('empresa_nome', '') or ''
                foto_path = colab.get('foto_path', '')
                status_detalhado = colab.get('status_detalhado', colab.get('status', ''))
                na_blocklist = colab.get('na_blocklist')

                # Definir cor do status
                if status_detalhado == 'Ativo':
                    cor_status = COR_SUCESSO
                elif status_detalhado == 'Block-List' or na_blocklist:
                    cor_status = COR_ERRO
                    status_detalhado = 'Block-List'
                else:
                    cor_status = COR_ALERTA

                # Widget de foto
                foto_widget = self._criar_avatar_foto(foto_path, cor_status)

                # Badge de status
                badge_status = ft.Container(
                    content=ft.Text(status_detalhado, color="white", size=10, weight=ft.FontWeight.BOLD),
                    bgcolor=cor_status,
                    padding=ft.padding.symmetric(horizontal=8, vertical=3),
                    border_radius=4,
                )

                self.lista_banco_talentos.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Container(content=foto_widget, width=50, alignment=ft.alignment.center),
                            ft.Container(content=ft.Text(colab.get('nome_completo', ''), size=14), expand=2),
                            ft.Container(content=ft.Text(formatar_cpf(colab.get('cpf', '')), size=14), expand=1),
                            ft.Container(content=ft.Text(colab.get('funcao', ''), size=14), expand=1),
                            ft.Container(content=ft.Text(empresa_nome, size=14), expand=2),
                            ft.Container(content=badge_status, width=90, alignment=ft.alignment.center),
                            ft.Container(
                                content=ft.IconButton(
                                    icon=ft.Icons.VISIBILITY,
                                    icon_color=COR_SECUNDARIA,
                                    tooltip="Ver ficha",
                                    on_click=lambda e, c=colab: self._abrir_ficha(c),
                                ),
                                width=50,
                            ),
                        ], spacing=15),
                        padding=ft.padding.symmetric(horizontal=20, vertical=8),
                        bgcolor=ft.Colors.WHITE,
                        border=ft.border.only(bottom=ft.BorderSide(1, COR_CINZA_CLARO)),
                        on_click=lambda e, c=colab: self._abrir_ficha(c),
                        ink=True,
                    )
                )

    def _filtrar_banco_talentos(self, e, tipo_filtro: str):
        """Aplica filtros no banco de talentos."""
        if tipo_filtro == 'empresa':
            self.filtro_empresa_banco = e.control.value if e.control.value else None
        elif tipo_filtro == 'status':
            self.filtro_status_banco = e.control.value if e.control.value else None
        elif tipo_filtro == 'texto':
            pass  # Será passado diretamente

        filtro_texto = self.campo_pesquisa_banco.value if hasattr(self, 'campo_pesquisa_banco') else None
        self._carregar_banco_talentos(filtro_texto)
        self.page.update()

    def _view_blocklist(self):
        # Usar a nova função que agrupa por CPF
        registros_agrupados = db.listar_blocklist_agrupado()
        lista = []

        def atualizar_lista():
            """Recarrega a lista de blocklist."""
            self.container_principal.content = self._view_blocklist()
            self.page.update()

        def editar_justificativa(entrada):
            """Abre o dialog para editar a justificativa de uma entrada específica."""
            campo_justificativa = ft.TextField(
                label="Justificativa",
                multiline=True,
                min_lines=3,
                max_lines=5,
                value=entrada.get('observacoes', ''),
                width=380,
            )

            def salvar(ev):
                # Verificar permissão
                if not db.usuario_pode_editar():
                    self.page.snack_bar = ft.SnackBar(
                        content=ft.Text("Você não tem permissão para editar dados."),
                        bgcolor=COR_ERRO
                    )
                    self.page.snack_bar.open = True
                    self.page.update()
                    return

                if not campo_justificativa.value or not campo_justificativa.value.strip():
                    self.page.snack_bar = ft.SnackBar(
                        content=ft.Text("Informe uma justificativa!"),
                        bgcolor=COR_ERRO
                    )
                    self.page.snack_bar.open = True
                    self.page.update()
                    return

                db.atualizar_blocklist(entrada['id'], campo_justificativa.value.strip())
                dialog.open = False
                self.page.snack_bar = ft.SnackBar(
                    content=ft.Text("Justificativa atualizada com sucesso!"),
                    bgcolor=COR_SUCESSO
                )
                self.page.snack_bar.open = True
                atualizar_lista()

            def cancelar(ev):
                dialog.open = False
                self.page.update()

            dialog = ft.AlertDialog(
                modal=True,
                title=ft.Row([
                    ft.Icon(ft.Icons.EDIT, color=COR_PRIMARIA, size=28),
                    ft.Text(f"Editar Justificativa - {entrada.get('entrada_numero', 1)}ª Entrada"),
                ]),
                content=ft.Container(
                    content=ft.Column([
                        ft.Text(f"Desligamento: {formatar_data_br(entrada.get('data_desligamento', ''))}", size=13),
                        ft.Text(f"Motivo: {entrada.get('motivo_desligamento', 'Não informado')}", size=13, color=ft.Colors.GREY_700),
                        ft.Divider(height=15),
                        campo_justificativa,
                    ], spacing=8),
                    width=400,
                ),
                actions=[
                    ft.TextButton("Cancelar", on_click=cancelar),
                    ft.ElevatedButton("Salvar", on_click=salvar, bgcolor=COR_SUCESSO, color="white"),
                ],
            )

            self.page.overlay.append(dialog)
            dialog.open = True
            self.page.update()

        def confirmar_remocao(entrada, nome_colaborador):
            """Abre o dialog para confirmar remoção de uma entrada da blocklist."""
            def confirmar(ev):
                # Verificar permissão
                if not db.usuario_pode_editar():
                    self.page.snack_bar = ft.SnackBar(
                        content=ft.Text("Você não tem permissão para remover dados."),
                        bgcolor=COR_ERRO
                    )
                    self.page.snack_bar.open = True
                    self.page.update()
                    return

                db.remover_blocklist(entrada['id'])
                dialog.open = False
                self.page.snack_bar = ft.SnackBar(
                    content=ft.Text(f"Entrada {entrada.get('entrada_numero', '')} de {nome_colaborador} removida!"),
                    bgcolor=COR_SUCESSO
                )
                self.page.snack_bar.open = True
                atualizar_lista()

            def cancelar(ev):
                dialog.open = False
                self.page.update()

            dialog = ft.AlertDialog(
                modal=True,
                title=ft.Row([
                    ft.Icon(ft.Icons.DELETE_FOREVER, color=COR_ERRO, size=28),
                    ft.Text("Remover Entrada da Block-List"),
                ]),
                content=ft.Column([
                    ft.Text(f"Deseja remover a {entrada.get('entrada_numero', 1)}ª entrada de:", size=14),
                    ft.Container(height=5),
                    ft.Text(nome_colaborador, weight=ft.FontWeight.BOLD, size=16),
                    ft.Container(height=10),
                    ft.Text(f"Desligamento: {formatar_data_br(entrada.get('data_desligamento', ''))}", size=13),
                    ft.Text(f"Motivo: {entrada.get('motivo_desligamento', 'Não informado')}", size=13),
                    ft.Container(height=10),
                    ft.Container(
                        content=ft.Text(
                            "Atenção: Esta ação não pode ser desfeita!",
                            color=COR_ERRO,
                            size=12,
                            italic=True,
                        ),
                    ),
                ], spacing=3),
                actions=[
                    ft.TextButton("Cancelar", on_click=cancelar),
                    ft.ElevatedButton("Remover", on_click=confirmar, bgcolor=COR_ERRO, color="white"),
                ],
            )

            self.page.overlay.append(dialog)
            dialog.open = True
            self.page.update()

        def mostrar_historico_completo(colaborador):
            """Mostra o histórico completo de um colaborador na Block-List."""
            nome = colaborador.get('nome', '')
            cpf = colaborador.get('cpf', '')
            historico = colaborador.get('historico', [])
            total = colaborador.get('total_entradas', 0)

            def fechar(ev):
                dialog.open = False
                self.page.update()

            # Criar lista de entradas do histórico
            lista_historico = []
            for entrada in historico:
                num = entrada.get('entrada_numero', 1)
                empresa = entrada.get('empresa_nome', 'Não informada')
                data_adm = formatar_data_br(entrada.get('data_admissao', ''))
                data_desl = formatar_data_br(entrada.get('data_desligamento', ''))
                motivo = entrada.get('motivo_desligamento', 'Não informado')
                obs = entrada.get('observacoes', '')

                # Cor do badge baseada no número da entrada
                if num == 1:
                    cor_badge = COR_ALERTA
                elif num == 2:
                    cor_badge = COR_ERRO
                else:
                    cor_badge = "#8B0000"  # Vermelho escuro para 3+ entradas

                # Construir conteúdo da entrada
                conteudo_entrada = [
                    # Cabeçalho com badge e botões
                    ft.Row([
                        ft.Container(
                            content=ft.Text(f"{num}ª Entrada", color="white", size=11, weight=ft.FontWeight.BOLD),
                            bgcolor=cor_badge,
                            padding=ft.padding.symmetric(horizontal=10, vertical=4),
                            border_radius=4,
                        ),
                        ft.Row([
                            ft.IconButton(
                                icon=ft.Icons.EDIT,
                                icon_size=16,
                                icon_color=COR_PRIMARIA,
                                tooltip="Editar justificativa",
                                on_click=lambda e, ent=entrada: (setattr(dialog, 'open', False), self.page.update(), editar_justificativa(ent)),
                            ),
                            ft.IconButton(
                                icon=ft.Icons.DELETE,
                                icon_size=16,
                                icon_color=COR_ERRO,
                                tooltip="Remover entrada",
                                on_click=lambda e, ent=entrada, n=nome: (setattr(dialog, 'open', False), self.page.update(), confirmar_remocao(ent, n)),
                            ),
                        ], spacing=0),
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                    # Empresa
                    ft.Row([
                        ft.Icon(ft.Icons.BUSINESS, size=14, color=ft.Colors.GREY_600),
                        ft.Text(empresa, size=12),
                    ], spacing=5),
                    # Datas
                    ft.Row([
                        ft.Icon(ft.Icons.DATE_RANGE, size=14, color=ft.Colors.GREY_600),
                        ft.Text(f"Admissão: {data_adm} | Desligamento: {data_desl}", size=11),
                    ], spacing=5),
                    # Motivo
                    ft.Row([
                        ft.Icon(ft.Icons.INFO_OUTLINE, size=14, color=ft.Colors.GREY_600),
                        ft.Text(f"Motivo: {motivo}", size=11),
                    ], spacing=5),
                ]

                # Adicionar justificativa (sempre mostrar, seja com conteúdo ou "Sem justificativa")
                if obs and obs.strip():
                    conteudo_entrada.append(
                        ft.Container(
                            content=ft.Text(obs, size=11, color=ft.Colors.GREY_700),
                            bgcolor=ft.Colors.GREY_100,
                            padding=8,
                            border_radius=4,
                        )
                    )
                else:
                    conteudo_entrada.append(
                        ft.Container(
                            content=ft.Text("Sem justificativa", size=11, italic=True, color=ft.Colors.GREY_500),
                            bgcolor=ft.Colors.GREY_100,
                            padding=8,
                            border_radius=4,
                        )
                    )

                lista_historico.append(
                    ft.Container(
                        content=ft.Column(conteudo_entrada, spacing=6),
                        padding=12,
                        border=ft.border.all(1, cor_badge),
                        border_radius=8,
                        bgcolor="white",
                    )
                )

            dialog = ft.AlertDialog(
                modal=True,
                title=ft.Row([
                    ft.Icon(ft.Icons.HISTORY, color=COR_ERRO, size=28),
                    ft.Column([
                        ft.Text("Histórico - Block-List", size=18),
                        ft.Text(f"{total} entrada(s) registrada(s)", size=12, color=ft.Colors.GREY_600),
                    ], spacing=0),
                ]),
                content=ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Text(nome, weight=ft.FontWeight.BOLD, size=16),
                        ]),
                        ft.Text(f"CPF: {formatar_cpf(cpf)}", size=13, color=ft.Colors.GREY_700),
                        ft.Divider(height=15),
                        ft.Container(
                            content=ft.Column(lista_historico, scroll=ft.ScrollMode.AUTO, spacing=10),
                            height=550,
                        ),
                    ], spacing=8),
                    width=650,
                ),
                actions=[
                    ft.ElevatedButton("Fechar", on_click=fechar, bgcolor=COR_SECUNDARIA, color="white"),
                ],
            )

            self.page.overlay.append(dialog)
            dialog.open = True
            self.page.update()

        # Criar cards para cada colaborador agrupado
        for colab in registros_agrupados:
            nome = colab.get('nome', '')
            cpf = colab.get('cpf', '')
            total = colab.get('total_entradas', 0)
            ultima = colab.get('ultima_entrada', {})
            motivo_ultima = ultima.get('motivo_desligamento', 'Não informado')
            data_ultima = formatar_data_br(ultima.get('data_desligamento', ''))

            # Definir cor e texto baseado no número de entradas
            if total == 1:
                cor_badge = COR_ALERTA
                texto_badge = "1ª vez"
            elif total == 2:
                cor_badge = COR_ERRO
                texto_badge = "2ª vez"
            else:
                cor_badge = "#8B0000"  # Vermelho escuro
                texto_badge = f"{total}ª vez"

            lista.append(
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Column([
                                ft.Text(nome, weight=ft.FontWeight.BOLD, size=15),
                                ft.Text(f"CPF: {formatar_cpf(cpf)}", size=12, color=ft.Colors.GREY_700),
                            ], spacing=2, expand=True),
                            ft.Container(
                                content=ft.Text(texto_badge, color="white", size=12, weight=ft.FontWeight.BOLD),
                                bgcolor=cor_badge,
                                padding=ft.padding.symmetric(horizontal=12, vertical=6),
                                border_radius=6,
                            ),
                        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                        ft.Divider(height=10, color=ft.Colors.GREY_300),
                        ft.Row([
                            ft.Icon(ft.Icons.EVENT, size=14, color=ft.Colors.GREY_600),
                            ft.Text(f"Último desligamento: {data_ultima}", size=12),
                        ], spacing=5),
                        ft.Row([
                            ft.Icon(ft.Icons.INFO_OUTLINE, size=14, color=ft.Colors.GREY_600),
                            ft.Text(f"Motivo: {motivo_ultima}", size=12),
                        ], spacing=5),
                        ft.Container(height=5),
                        ft.Row([
                            ft.Icon(ft.Icons.HISTORY, size=14, color=COR_SECUNDARIA),
                            ft.Text(
                                f"Clique para ver histórico completo ({total} entrada{'s' if total > 1 else ''})",
                                size=11,
                                italic=True,
                                color=COR_SECUNDARIA,
                            ),
                        ], spacing=5),
                    ], spacing=5),
                    padding=15,
                    border=ft.border.all(1, cor_badge),
                    border_radius=8,
                    bgcolor="white",
                    on_click=lambda e, c=colab: mostrar_historico_completo(c),
                    ink=True,
                )
            )

        # Estatísticas
        total_pessoas = len(registros_agrupados)
        reincidentes = len([c for c in registros_agrupados if c.get('total_entradas', 0) > 1])

        return ft.Column([
            ft.Container(
                content=ft.Row([
                    ft.Row([
                        ft.Icon(ft.Icons.BLOCK, size=30, color=COR_ERRO),
                        ft.Text("Block-List (Ex-Colaboradores)", size=24, weight=ft.FontWeight.BOLD),
                    ], spacing=10),
                    ft.Row([
                        ft.Container(
                            content=ft.Text(f"{total_pessoas} pessoa(s)", color="white", size=12),
                            bgcolor=COR_SECUNDARIA,
                            padding=ft.padding.symmetric(horizontal=10, vertical=5),
                            border_radius=4,
                        ),
                        ft.Container(
                            content=ft.Text(f"{reincidentes} reincidente(s)", color="white", size=12),
                            bgcolor=COR_ERRO if reincidentes > 0 else ft.Colors.GREY,
                            padding=ft.padding.symmetric(horizontal=10, vertical=5),
                            border_radius=4,
                            visible=True,
                        ),
                    ], spacing=10),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                padding=15, bgcolor="white", border_radius=8,
            ),
            ft.Container(
                content=ft.Text(
                    "Cada colaborador aparece apenas uma vez. Clique para ver o histórico completo de entradas.",
                    size=12,
                    italic=True,
                    color=ft.Colors.GREY_700,
                ),
                padding=ft.padding.only(left=10, top=5),
            ),
            ft.Container(
                content=ft.Column(lista if lista else [
                    ft.Container(
                        content=ft.Column([
                            ft.Icon(ft.Icons.CHECK_CIRCLE, size=50, color=COR_SUCESSO),
                            ft.Text("Nenhum registro na Block-List!", size=16, color=COR_SUCESSO),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                        padding=30,
                    )
                ], scroll=ft.ScrollMode.AUTO, spacing=10),
                expand=True, padding=10,
            ),
        ], spacing=10, expand=True)
    
    def _view_empresas(self):
        empresas = db.listar_empresas(apenas_ativas=False)
        self.form_empresa_container = ft.Container()

        lista = []
        for emp in empresas:
            emp_id = emp.get('id')
            qtd_colab = db.contar_colaboradores_empresa(emp_id)

            lista.append(
                ft.Container(
                    content=ft.Row([
                        ft.Column([
                            ft.Text(emp.get('razao_social', ''), weight=ft.FontWeight.BOLD),
                            ft.Text(f"CNPJ: {emp.get('cnpj', '')}", size=13),
                            ft.Row([
                                ft.Text(f"{emp.get('cidade', '')} - {emp.get('uf', '')}", size=12, color=ft.Colors.GREY),
                                ft.Text(f"| {qtd_colab} colaborador(es)", size=12, color=COR_SECUNDARIA),
                            ], spacing=5),
                        ], spacing=2, expand=True),
                        ft.Row([
                            ft.Container(
                                content=ft.Text("Ativa" if emp.get('ativa') else "Inativa", color="white", size=11),
                                bgcolor=COR_SUCESSO if emp.get('ativa') else COR_ERRO,
                                padding=ft.padding.symmetric(horizontal=10, vertical=3), border_radius=4,
                            ),
                            ft.IconButton(
                                icon=ft.Icons.DELETE,
                                icon_color=COR_ERRO,
                                tooltip="Excluir empresa",
                                on_click=lambda e, eid=emp_id, nome=emp.get('razao_social', ''), qtd=qtd_colab: self._confirmar_exclusao_empresa(eid, nome, qtd),
                            ),
                        ], spacing=5),
                    ]),
                    padding=15, border=ft.border.all(1, COR_SECUNDARIA), border_radius=8, bgcolor="white",
                )
            )

        return ft.Column([
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.BUSINESS, size=30, color=COR_PRIMARIA),
                    ft.Text("Empresas Contratantes", size=24, weight=ft.FontWeight.BOLD),
                    ft.ElevatedButton("Nova Empresa", icon=ft.Icons.ADD, on_click=self._mostrar_form_empresa,
                                     bgcolor=COR_SUCESSO, color="white"),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                padding=15, bgcolor="white", border_radius=8,
            ),
            self.form_empresa_container,
            ft.Container(
                content=ft.Column(lista, scroll=ft.ScrollMode.AUTO, spacing=10),
                expand=True, padding=10,
            ),
        ], spacing=10, expand=True)

    def _view_dashboard(self):
        """Constrói a view do Dashboard com gráficos e estatísticas."""
        dashboard = DashboardView(self.page)
        return dashboard.build()

    def _mostrar_form_empresa(self, e):
        self.razao_social = ft.TextField(label="Razão Social *", width=400)
        self.cnpj_empresa = ft.TextField(label="CNPJ *", width=200)
        self.cidade_empresa = ft.TextField(label="Cidade", width=200)
        self.uf_empresa = ft.Dropdown(label="UF", width=80,
                                       options=[ft.dropdown.Option(uf) for uf in ['AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA', 'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN', 'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO']])
        
        self.form_empresa_container.content = ft.Container(
            content=ft.Column([
                ft.Text("Nova Empresa", size=16, weight=ft.FontWeight.BOLD),
                ft.Divider(),
                ft.Row([self.razao_social, self.cnpj_empresa], wrap=True),
                ft.Row([self.cidade_empresa, self.uf_empresa], wrap=True),
                ft.Row([
                    ft.ElevatedButton("Salvar", icon=ft.Icons.SAVE, on_click=self._salvar_empresa,
                                     bgcolor=COR_SUCESSO, color="white"),
                    ft.OutlinedButton("Cancelar", on_click=lambda e: self._cancelar_form_empresa()),
                ], spacing=10),
            ], spacing=10),
            padding=15, border=ft.border.all(1, COR_SECUNDARIA), border_radius=8, bgcolor="white",
        )
        self.page.update()
    
    def _salvar_empresa(self, e):
        if not self.razao_social.value or not self.cnpj_empresa.value:
            self.page.snack_bar = ft.SnackBar(content=ft.Text("Preencha os campos obrigatórios!"), bgcolor=COR_ERRO)
            self.page.snack_bar.open = True
            self.page.update()
            return
        
        db.criar_empresa({
            'razao_social': self.razao_social.value,
            'cnpj': self.cnpj_empresa.value,
            'cidade': self.cidade_empresa.value,
            'uf': self.uf_empresa.value,
        })
        
        self._cancelar_form_empresa()
        self.navegar("empresas")
        
        self.page.snack_bar = ft.SnackBar(content=ft.Text("Empresa cadastrada!"), bgcolor=COR_SUCESSO)
        self.page.snack_bar.open = True
        self.page.update()
    
    def _cancelar_form_empresa(self):
        self.form_empresa_container.content = None
        self.page.update()

    def _confirmar_exclusao_empresa(self, empresa_id: int, nome_empresa: str, qtd_colaboradores: int):
        """Mostra diálogo de confirmação para exclusão de empresa."""

        def fechar_dialog(e=None):
            dialog.open = False
            self.page.update()

        def excluir_empresa(e=None, excluir_colaboradores: bool = False):
            try:
                sucesso = db.excluir_empresa(empresa_id, excluir_colaboradores=excluir_colaboradores)
                fechar_dialog()

                if sucesso:
                    msg = f"Empresa '{nome_empresa}' excluída com sucesso!"
                    if excluir_colaboradores and qtd_colaboradores > 0:
                        msg += f" ({qtd_colaboradores} colaborador(es) também foram excluídos)"
                    self.page.snack_bar = ft.SnackBar(content=ft.Text(msg), bgcolor=COR_SUCESSO)
                else:
                    self.page.snack_bar = ft.SnackBar(
                        content=ft.Text("Não foi possível excluir. Há colaboradores vinculados."),
                        bgcolor=COR_ERRO
                    )
                self.page.snack_bar.open = True
                self.navegar("empresas")
            except (ValueError, KeyError, TypeError) as ex:
                fechar_dialog()
                db.registrar_log("sistema", "erro", f"Erro ao excluir empresa: {str(ex)}")
                self.page.snack_bar = ft.SnackBar(content=ft.Text(f"Erro ao excluir: {str(ex)}"), bgcolor=COR_ERRO)
                self.page.snack_bar.open = True
                self.page.update()

        # Criar conteúdo do diálogo baseado na quantidade de colaboradores
        if qtd_colaboradores > 0:
            conteudo = ft.Column([
                ft.Text(f"A empresa '{nome_empresa}' possui {qtd_colaboradores} colaborador(es) cadastrado(s)."),
                ft.Container(height=10),
                ft.Text("O que deseja fazer?", weight=ft.FontWeight.BOLD),
                ft.Container(height=5),
                ft.Text("• Excluir apenas a empresa: Os colaboradores ficarão sem empresa vinculada", size=13),
                ft.Text("• Excluir tudo: A empresa E todos os colaboradores serão excluídos permanentemente", size=13, color=COR_ERRO),
            ], spacing=5)

            acoes = [
                ft.TextButton("Cancelar", on_click=fechar_dialog),
                ft.ElevatedButton(
                    "Excluir Tudo",
                    icon=ft.Icons.DELETE_FOREVER,
                    bgcolor=COR_ERRO,
                    color="white",
                    on_click=lambda e: excluir_empresa(excluir_colaboradores=True),
                ),
            ]
        else:
            conteudo = ft.Text(f"Tem certeza que deseja excluir a empresa '{nome_empresa}'?")
            acoes = [
                ft.TextButton("Cancelar", on_click=fechar_dialog),
                ft.ElevatedButton(
                    "Excluir",
                    icon=ft.Icons.DELETE,
                    bgcolor=COR_ERRO,
                    color="white",
                    on_click=lambda e: excluir_empresa(excluir_colaboradores=False),
                ),
            ]

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.WARNING, color=COR_ALERTA, size=30),
                ft.Text("Confirmar Exclusão", color=COR_ALERTA),
            ]),
            content=ft.Container(content=conteudo, width=450),
            actions=acoes,
            actions_alignment=ft.MainAxisAlignment.END,
        )

        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def _exportar_excel(self):
        """Exibe diálogo para escolher empresa e exportar Excel completo."""
        from utilities.excel_export import exportar_completo_excel

        # Carregar empresas para o dropdown
        empresas = db.listar_empresas()

        empresa_selecionada = {"id": None, "nome": "Todas as Empresas"}

        def fechar(ev=None):
            dialog.open = False
            self.page.update()

        def on_empresa_change(e):
            if e.control.value == "todas":
                empresa_selecionada["id"] = None
                empresa_selecionada["nome"] = "Todas as Empresas"
            else:
                empresa_id = int(e.control.value)
                empresa_selecionada["id"] = empresa_id
                for emp in empresas:
                    if emp['id'] == empresa_id:
                        empresa_selecionada["nome"] = emp['razao_social']
                        break

        def exportar(ev):
            """Inicia o processo de exportação."""

            def on_resultado_salvar(e: ft.FilePickerResultEvent):
                if e.path:
                    try:
                        # Coletar todos os dados
                        empresa_id = empresa_selecionada["id"]

                        # Colaboradores ativos
                        colaboradores_ativos = db.listar_colaboradores(
                            status='ATIVO',
                            empresa_id=empresa_id
                        )

                        # Colaboradores inativos
                        colaboradores_inativos = db.listar_colaboradores(
                            status='INATIVO',
                            empresa_id=empresa_id
                        )

                        # Contratos
                        contratos = db.listar_todos_contratos_com_colaborador(empresa_id)

                        # Férias
                        ferias = db.listar_todas_ferias_com_colaborador(empresa_id)

                        # Dependentes
                        dependentes = db.listar_todos_dependentes_com_colaborador(empresa_id)

                        # Block-list (sempre completa)
                        blocklist = db.listar_blocklist_completa()

                        # Documentos pendentes
                        docs_pendentes = db.listar_documentos_pendentes_todos(empresa_id)

                        # Exportar
                        output = exportar_completo_excel(
                            colaboradores_ativos=colaboradores_ativos,
                            colaboradores_inativos=colaboradores_inativos,
                            contratos=contratos,
                            ferias=ferias,
                            dependentes=dependentes,
                            blocklist=blocklist,
                            documentos_pendentes=docs_pendentes,
                            output_path=e.path,
                            empresa_nome=empresa_selecionada["nome"]
                        )

                        # Registrar log
                        db.registrar_log(
                            tipo_acao='EXPORTAR',
                            categoria='EXCEL',
                            descricao=f'Relatório Excel exportado - Empresa: {empresa_selecionada["nome"]}',
                            entidade_tipo='excel',
                            entidade_nome=os.path.basename(e.path)
                        )

                        self.page.snack_bar = ft.SnackBar(
                            content=ft.Text(f"Excel exportado: {e.path}"),
                            bgcolor=COR_SUCESSO
                        )
                    except (IOError, OSError, PermissionError) as ex:
                        db.registrar_log("sistema", "erro", f"Erro ao exportar Excel: {str(ex)}")
                        self.page.snack_bar = ft.SnackBar(
                            content=ft.Text(f"Erro ao exportar: {str(ex)}"),
                            bgcolor=COR_ERRO
                        )
                    self.page.snack_bar.open = True
                    self.page.update()

                # Fechar diálogo
                fechar()

            # Criar file picker para salvar
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_picker = ft.FilePicker(on_result=on_resultado_salvar)
            self.page.overlay.append(file_picker)
            self.page.update()
            file_picker.save_file(
                dialog_title="Salvar Relatório Excel",
                file_name=f"RH_RENOVO_{timestamp}.xlsx",
                allowed_extensions=["xlsx"]
            )

        # Criar dropdown de empresas
        opcoes_empresas = [ft.dropdown.Option(key="todas", text="Todas as Empresas")]
        for emp in empresas:
            opcoes_empresas.append(
                ft.dropdown.Option(key=str(emp['id']), text=emp['razao_social'])
            )

        dropdown_empresa = ft.Dropdown(
            label="Selecione a Empresa",
            options=opcoes_empresas,
            value="todas",
            on_change=on_empresa_change,
            border_color=COR_SECUNDARIA,
            width=350,
        )

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.DOWNLOAD, color=COR_PRIMARIA, size=28),
                ft.Text("Exportar Excel"),
            ]),
            content=ft.Container(
                content=ft.Column([
                    ft.Text(
                        "O relatório será exportado com as seguintes abas:",
                        size=12,
                        color=ft.Colors.GREY_700
                    ),
                    ft.Container(height=10),

                    # Lista de abas
                    ft.Container(
                        content=ft.Column([
                            ft.Row([ft.Icon(ft.Icons.CHECK, size=14, color=COR_SUCESSO), ft.Text("Colaboradores Ativos", size=11)]),
                            ft.Row([ft.Icon(ft.Icons.CHECK, size=14, color=COR_SUCESSO), ft.Text("Colaboradores Inativos", size=11)]),
                            ft.Row([ft.Icon(ft.Icons.CHECK, size=14, color=COR_SUCESSO), ft.Text("Contratos de Experiência", size=11)]),
                            ft.Row([ft.Icon(ft.Icons.CHECK, size=14, color=COR_SUCESSO), ft.Text("Férias", size=11)]),
                            ft.Row([ft.Icon(ft.Icons.CHECK, size=14, color=COR_SUCESSO), ft.Text("Dependentes", size=11)]),
                            ft.Row([ft.Icon(ft.Icons.CHECK, size=14, color=COR_SUCESSO), ft.Text("Block-List (completa)", size=11)]),
                            ft.Row([ft.Icon(ft.Icons.CHECK, size=14, color=COR_SUCESSO), ft.Text("Documentos Pendentes", size=11)]),
                        ], spacing=5),
                        padding=10,
                        bgcolor=ft.Colors.GREY_100,
                        border_radius=8,
                    ),

                    ft.Container(height=15),

                    # Dropdown de empresa
                    dropdown_empresa,

                ], spacing=0),
                width=380,
                height=280,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=fechar),
                ft.ElevatedButton(
                    "Exportar",
                    on_click=exportar,
                    bgcolor=COR_SUCESSO,
                    color="white",
                    icon=ft.Icons.DOWNLOAD
                ),
            ],
        )

        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()

    def _fazer_backup(self):
        """Exibe diálogo para escolher entre Importar ou Exportar backup."""

        def fechar(ev=None):
            dialog.open = False
            self.page.update()

        def exportar_backup(ev):
            """Exporta o banco de dados para um local escolhido pelo usuário."""

            def on_resultado_exportar(e: ft.FilePickerResultEvent):
                if e.path:
                    try:
                        # Copiar banco de dados para o local escolhido
                        shutil.copy2(db.DATABASE_PATH, e.path)

                        # Registrar log
                        db.registrar_log(
                            tipo_acao='EXPORTAR',
                            categoria='BACKUP',
                            descricao=f'Backup exportado para: {e.path}',
                            entidade_tipo='backup',
                            entidade_nome=os.path.basename(e.path)
                        )

                        self.page.snack_bar = ft.SnackBar(
                            content=ft.Text(f"Backup exportado: {e.path}"),
                            bgcolor=COR_SUCESSO
                        )
                    except (IOError, OSError, PermissionError) as ex:
                        db.registrar_log("sistema", "erro", f"Erro ao exportar backup: {str(ex)}")
                        self.page.snack_bar = ft.SnackBar(
                            content=ft.Text(f"Erro ao exportar: {str(ex)}"),
                            bgcolor=COR_ERRO
                        )
                    self.page.snack_bar.open = True
                    self.page.update()
                # Fechar diálogo após escolher ou cancelar
                fechar()

            # Criar file picker para salvar
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_picker = ft.FilePicker(on_result=on_resultado_exportar)
            self.page.overlay.append(file_picker)
            self.page.update()
            file_picker.save_file(
                dialog_title="Exportar Backup",
                file_name=f"rh_backup_{timestamp}.db",
                allowed_extensions=["db"]
            )

        def importar_backup(ev):
            """Importa um banco de dados de um arquivo escolhido."""

            def on_resultado_importar(e: ft.FilePickerResultEvent):
                if e.files and len(e.files) > 0:
                    arquivo_selecionado = e.files[0].path

                    # Validar se é um arquivo .db
                    if not arquivo_selecionado.lower().endswith('.db'):
                        self.page.snack_bar = ft.SnackBar(
                            content=ft.Text("Arquivo inválido. Selecione um arquivo .db"),
                            bgcolor=COR_ERRO
                        )
                        self.page.snack_bar.open = True
                        self.page.update()
                        return

                    # Fechar diálogo principal e mostrar confirmação
                    fechar()
                    mostrar_confirmacao_importar(arquivo_selecionado)
                else:
                    # Usuário cancelou - não fecha o diálogo
                    pass

            file_picker = ft.FilePicker(on_result=on_resultado_importar)
            self.page.overlay.append(file_picker)
            self.page.update()
            file_picker.pick_files(
                dialog_title="Selecionar Backup para Importar",
                allowed_extensions=["db"],
                allow_multiple=False
            )

        def mostrar_confirmacao_importar(arquivo_path):
            """Mostra diálogo de confirmação antes de importar."""

            def cancelar(ev):
                dialog_confirm.open = False
                self.page.update()

            def confirmar_importacao(ev):
                dialog_confirm.open = False
                self.page.update()

                try:
                    # Validar se o arquivo é um banco de dados válido do sistema
                    if not db.validar_banco_dados(arquivo_path):
                        self.page.snack_bar = ft.SnackBar(
                            content=ft.Text("Arquivo inválido. Não é um banco de dados válido do sistema."),
                            bgcolor=COR_ERRO
                        )
                        self.page.snack_bar.open = True
                        self.page.update()
                        return

                    # Fazer backup do banco atual antes de importar
                    backup_atual = db.backup_database()

                    # Importar o novo banco
                    shutil.copy2(arquivo_path, db.DATABASE_PATH)

                    # Registrar log
                    db.registrar_log(
                        tipo_acao='IMPORTAR',
                        categoria='BACKUP',
                        descricao=f'Backup importado de: {arquivo_path}',
                        entidade_tipo='backup',
                        entidade_nome=os.path.basename(arquivo_path)
                    )

                    # Mostrar mensagem de sucesso e reiniciar
                    self.page.snack_bar = ft.SnackBar(
                        content=ft.Text("Backup importado com sucesso! O programa será reiniciado."),
                        bgcolor=COR_SUCESSO
                    )
                    self.page.snack_bar.open = True
                    self.page.update()

                    # Aguardar um pouco e reiniciar
                    time.sleep(2)
                    self.page.window.close()

                except (IOError, OSError, PermissionError) as ex:
                    db.registrar_log("sistema", "erro", f"Erro ao importar backup: {str(ex)}")
                    self.page.snack_bar = ft.SnackBar(
                        content=ft.Text(f"Erro ao importar: {str(ex)}"),
                        bgcolor=COR_ERRO
                    )
                    self.page.snack_bar.open = True
                    self.page.update()

            dialog_confirm = ft.AlertDialog(
                modal=True,
                title=ft.Row([
                    ft.Icon(ft.Icons.WARNING, color=COR_ALERTA, size=28),
                    ft.Text("Confirmar Importação"),
                ]),
                content=ft.Container(
                    content=ft.Column([
                        ft.Text(
                            "ATENÇÃO: Esta ação irá substituir TODOS os dados atuais!",
                            weight=ft.FontWeight.BOLD,
                            color=COR_ERRO
                        ),
                        ft.Container(height=10),
                        ft.Text("• Um backup do banco atual será feito automaticamente"),
                        ft.Text("• Todos os colaboradores, empresas e documentos serão substituídos"),
                        ft.Text("• O programa será reiniciado após a importação"),
                        ft.Container(height=10),
                        ft.Text(f"Arquivo selecionado:", size=11, color=ft.Colors.GREY_600),
                        ft.Text(os.path.basename(arquivo_path), weight=ft.FontWeight.BOLD, size=12),
                        ft.Container(height=10),
                        ft.Text("Deseja continuar?", weight=ft.FontWeight.BOLD),
                    ], spacing=5),
                    width=400,
                ),
                actions=[
                    ft.TextButton("Cancelar", on_click=cancelar),
                    ft.ElevatedButton(
                        "Sim, Importar",
                        on_click=confirmar_importacao,
                        bgcolor=COR_ERRO,
                        color="white"
                    ),
                ],
            )

            self.page.overlay.append(dialog_confirm)
            dialog_confirm.open = True
            self.page.update()

        # Diálogo principal com opções Importar/Exportar
        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.BACKUP, color=COR_PRIMARIA, size=28),
                ft.Text("Backup do Sistema"),
            ]),
            content=ft.Container(
                content=ft.Column([
                    ft.Text(
                        "Escolha uma opção:",
                        size=14,
                        color=ft.Colors.GREY_700
                    ),
                    ft.Container(height=15),

                    # Botão Exportar
                    ft.Container(
                        content=ft.Row([
                            ft.Container(
                                content=ft.Icon(ft.Icons.UPLOAD_FILE, size=30, color="white"),
                                bgcolor=COR_SUCESSO,
                                padding=15,
                                border_radius=10,
                            ),
                            ft.Column([
                                ft.Text("Exportar Backup", weight=ft.FontWeight.BOLD, size=14),
                                ft.Text("Salvar uma cópia do banco de dados", size=11, color=ft.Colors.GREY_600),
                            ], spacing=2, expand=True),
                            ft.Icon(ft.Icons.CHEVRON_RIGHT, color=ft.Colors.GREY_400),
                        ], spacing=15),
                        padding=15,
                        border=ft.border.all(1, COR_CINZA_CLARO),
                        border_radius=10,
                        on_click=exportar_backup,
                        ink=True,
                    ),

                    ft.Container(height=10),

                    # Botão Importar
                    ft.Container(
                        content=ft.Row([
                            ft.Container(
                                content=ft.Icon(ft.Icons.DOWNLOAD, size=30, color="white"),
                                bgcolor=COR_PRIMARIA,
                                padding=15,
                                border_radius=10,
                            ),
                            ft.Column([
                                ft.Text("Importar Backup", weight=ft.FontWeight.BOLD, size=14),
                                ft.Text("Restaurar dados de um backup anterior", size=11, color=ft.Colors.GREY_600),
                            ], spacing=2, expand=True),
                            ft.Icon(ft.Icons.CHEVRON_RIGHT, color=ft.Colors.GREY_400),
                        ], spacing=15),
                        padding=15,
                        border=ft.border.all(1, COR_CINZA_CLARO),
                        border_radius=10,
                        on_click=importar_backup,
                        ink=True,
                    ),

                ], spacing=0),
                width=350,
                height=230,
            ),
            actions=[
                ft.TextButton("Fechar", on_click=fechar),
            ],
        )

        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()

    def _mostrar_log(self):
        """Exibe a tela de Log com histórico completo do sistema."""

        # Variáveis de estado para filtros e paginação
        self.log_filtro_categoria = None
        self.log_filtro_tipo_acao = None
        self.log_filtro_pesquisa = None
        self.log_pagina_atual = 0
        self.log_itens_por_pagina = 50

        def fechar(ev):
            dialog.open = False
            self.page.update()

        def aplicar_filtros(e=None):
            self.log_pagina_atual = 0
            atualizar_lista_logs()

        def limpar_filtros(e=None):
            dropdown_categoria.value = ""
            dropdown_tipo_acao.value = ""
            campo_pesquisa.value = ""
            self.log_filtro_categoria = None
            self.log_filtro_tipo_acao = None
            self.log_filtro_pesquisa = None
            self.log_pagina_atual = 0
            atualizar_lista_logs()
            self.page.update()

        def pagina_anterior(e):
            if self.log_pagina_atual > 0:
                self.log_pagina_atual -= 1
                atualizar_lista_logs()

        def proxima_pagina(e):
            total = db.contar_logs(
                categoria=self.log_filtro_categoria,
                tipo_acao=self.log_filtro_tipo_acao,
                pesquisa=self.log_filtro_pesquisa
            )
            max_paginas = (total // self.log_itens_por_pagina) + (1 if total % self.log_itens_por_pagina > 0 else 0)
            if self.log_pagina_atual < max_paginas - 1:
                self.log_pagina_atual += 1
                atualizar_lista_logs()

        def on_categoria_change(e):
            self.log_filtro_categoria = e.control.value if e.control.value else None
            aplicar_filtros()

        def on_tipo_acao_change(e):
            self.log_filtro_tipo_acao = e.control.value if e.control.value else None
            aplicar_filtros()

        def on_pesquisa_change(e):
            self.log_filtro_pesquisa = e.control.value if e.control.value and len(e.control.value) >= 2 else None
            if self.log_filtro_pesquisa or not e.control.value:
                aplicar_filtros()

        def formatar_data_hora(data_hora_str):
            """Formata data e hora para exibição."""
            if not data_hora_str:
                return "-"
            try:
                dt = datetime.strptime(data_hora_str, '%Y-%m-%d %H:%M:%S')
                return dt.strftime('%d/%m/%Y %H:%M:%S')
            except (ValueError, TypeError):
                try:
                    dt = datetime.fromisoformat(data_hora_str)
                    return dt.strftime('%d/%m/%Y %H:%M:%S')
                except (ValueError, TypeError):
                    return data_hora_str

        def obter_icone_categoria(categoria):
            """Retorna o ícone correspondente à categoria."""
            icones = {
                'COLABORADOR': ft.Icons.PERSON,
                'EMPRESA': ft.Icons.BUSINESS,
                'DOCUMENTO': ft.Icons.ATTACH_FILE,
                'FERIAS': ft.Icons.BEACH_ACCESS,
                'CONTRATO': ft.Icons.SCHEDULE,
                'BLOCKLIST': ft.Icons.BLOCK,
                'BACKUP': ft.Icons.BACKUP,
                'DEPENDENTE': ft.Icons.FAMILY_RESTROOM,
                'SISTEMA': ft.Icons.SETTINGS,
            }
            return icones.get(categoria, ft.Icons.INFO)

        def obter_cor_tipo_acao(tipo_acao):
            """Retorna a cor correspondente ao tipo de ação."""
            cores = {
                'CRIAR': COR_SUCESSO,
                'EDITAR': COR_PRIMARIA,
                'EXCLUIR': COR_ERRO,
                'ANEXAR': COR_SECUNDARIA,
                'DESATIVAR': COR_ALERTA,
                'REATIVAR': COR_SUCESSO,
                'PRORROGAR': COR_PRIMARIA,
                'CONVERTER': COR_SECUNDARIA,
                'REGISTRAR': COR_SECUNDARIA,
            }
            return cores.get(tipo_acao, ft.Colors.GREY)

        def atualizar_lista_logs():
            """Atualiza a lista de logs com os filtros aplicados."""
            logs = db.listar_logs(
                limite=self.log_itens_por_pagina,
                offset=self.log_pagina_atual * self.log_itens_por_pagina,
                categoria=self.log_filtro_categoria,
                tipo_acao=self.log_filtro_tipo_acao,
                pesquisa=self.log_filtro_pesquisa
            )

            total = db.contar_logs(
                categoria=self.log_filtro_categoria,
                tipo_acao=self.log_filtro_tipo_acao,
                pesquisa=self.log_filtro_pesquisa
            )

            lista_logs.controls.clear()

            if not logs:
                lista_logs.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Icon(ft.Icons.HISTORY, size=50, color=ft.Colors.GREY),
                            ft.Text("Nenhum registro encontrado", italic=True, color=ft.Colors.GREY),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                        padding=30,
                        alignment=ft.alignment.center,
                    )
                )
            else:
                for log in logs:
                    categoria = log.get('categoria', '')
                    tipo_acao = log.get('tipo_acao', '')
                    descricao = log.get('descricao', '')
                    entidade_nome = log.get('entidade_nome', '')
                    data_hora = formatar_data_hora(log.get('data_hora', ''))
                    usuario = log.get('usuario', 'Sistema')
                    valor_anterior = log.get('valor_anterior', '')
                    valor_novo = log.get('valor_novo', '')

                    # Construir conteúdo da linha
                    conteudo_linha = [
                        ft.Row([
                            ft.Container(
                                content=ft.Icon(obter_icone_categoria(categoria), size=16, color="white"),
                                bgcolor=obter_cor_tipo_acao(tipo_acao),
                                padding=6,
                                border_radius=4,
                            ),
                            ft.Container(
                                content=ft.Text(tipo_acao, color="white", size=10, weight=ft.FontWeight.BOLD),
                                bgcolor=obter_cor_tipo_acao(tipo_acao),
                                padding=ft.padding.symmetric(horizontal=8, vertical=3),
                                border_radius=4,
                            ),
                            ft.Container(
                                content=ft.Text(categoria, size=10, color=ft.Colors.GREY_700),
                                bgcolor=ft.Colors.GREY_200,
                                padding=ft.padding.symmetric(horizontal=6, vertical=2),
                                border_radius=4,
                            ),
                            ft.Text(data_hora, size=11, color=ft.Colors.GREY_600),
                        ], spacing=8),
                        ft.Text(descricao, size=12, weight=ft.FontWeight.W_500),
                    ]

                    # Adicionar nome da entidade se existir
                    if entidade_nome:
                        conteudo_linha.append(
                            ft.Row([
                                ft.Icon(ft.Icons.LABEL, size=12, color=ft.Colors.GREY_500),
                                ft.Text(entidade_nome, size=11, color=ft.Colors.GREY_700),
                            ], spacing=5)
                        )

                    # Adicionar valores alterados se existirem
                    if valor_anterior or valor_novo:
                        valores_row = ft.Row([
                            ft.Icon(ft.Icons.SWAP_HORIZ, size=12, color=ft.Colors.GREY_500),
                        ], spacing=5)
                        if valor_anterior:
                            valores_row.controls.append(
                                ft.Container(
                                    content=ft.Text(valor_anterior[:50] + ('...' if len(str(valor_anterior)) > 50 else ''),
                                                   size=10, color=COR_ERRO),
                                    bgcolor="#FFEBEE",
                                    padding=ft.padding.symmetric(horizontal=4, vertical=2),
                                    border_radius=3,
                                )
                            )
                        if valor_anterior and valor_novo:
                            valores_row.controls.append(ft.Text("→", size=10, color=ft.Colors.GREY))
                        if valor_novo:
                            valores_row.controls.append(
                                ft.Container(
                                    content=ft.Text(valor_novo[:50] + ('...' if len(str(valor_novo)) > 50 else ''),
                                                   size=10, color=COR_SUCESSO),
                                    bgcolor="#E8F5E9",
                                    padding=ft.padding.symmetric(horizontal=4, vertical=2),
                                    border_radius=3,
                                )
                            )
                        conteudo_linha.append(valores_row)

                    # Adicionar usuário
                    conteudo_linha.append(
                        ft.Row([
                            ft.Icon(ft.Icons.PERSON, size=12, color=ft.Colors.GREY_400),
                            ft.Text(usuario, size=10, color=ft.Colors.GREY_500, italic=True),
                        ], spacing=3)
                    )

                    lista_logs.controls.append(
                        ft.Container(
                            content=ft.Column(conteudo_linha, spacing=4),
                            padding=12,
                            border=ft.border.only(bottom=ft.BorderSide(1, COR_CINZA_CLARO)),
                            bgcolor="white",
                        )
                    )

            # Atualizar info de paginação
            max_paginas = (total // self.log_itens_por_pagina) + (1 if total % self.log_itens_por_pagina > 0 else 0)
            if max_paginas == 0:
                max_paginas = 1
            texto_paginacao.value = f"Página {self.log_pagina_atual + 1} de {max_paginas} ({total} registros)"
            btn_anterior.disabled = self.log_pagina_atual == 0
            btn_proximo.disabled = self.log_pagina_atual >= max_paginas - 1

            self.page.update()

        # Obter estatísticas
        stats = db.obter_estatisticas_log()

        # Obter categorias e tipos de ação existentes
        categorias_existentes = db.obter_categorias_log()
        tipos_acao_existentes = db.obter_tipos_acao_log()

        # Criar dropdowns de filtro
        dropdown_categoria = ft.Dropdown(
            label="Categoria",
            width=150,
            options=[ft.dropdown.Option(key="", text="Todas")] +
                    [ft.dropdown.Option(key=c, text=c) for c in categorias_existentes],
            value="",
            on_change=on_categoria_change,
            border_color=COR_SECUNDARIA,
            dense=True,
        )

        dropdown_tipo_acao = ft.Dropdown(
            label="Tipo de Ação",
            width=150,
            options=[ft.dropdown.Option(key="", text="Todos")] +
                    [ft.dropdown.Option(key=t, text=t) for t in tipos_acao_existentes],
            value="",
            on_change=on_tipo_acao_change,
            border_color=COR_SECUNDARIA,
            dense=True,
        )

        campo_pesquisa = ft.TextField(
            label="Pesquisar",
            prefix_icon=ft.Icons.SEARCH,
            width=200,
            on_change=on_pesquisa_change,
            border_color=COR_SECUNDARIA,
            dense=True,
        )

        # Lista de logs
        lista_logs = ft.Column(scroll=ft.ScrollMode.AUTO, spacing=0)

        # Controles de paginação
        texto_paginacao = ft.Text("", size=12, color=ft.Colors.GREY_700)
        btn_anterior = ft.IconButton(
            icon=ft.Icons.CHEVRON_LEFT,
            on_click=pagina_anterior,
            icon_color=COR_PRIMARIA,
        )
        btn_proximo = ft.IconButton(
            icon=ft.Icons.CHEVRON_RIGHT,
            on_click=proxima_pagina,
            icon_color=COR_PRIMARIA,
        )

        # Cards de estatísticas
        cards_stats = ft.Row([
            self._criar_card_stat("Total", stats['total'], ft.Icons.HISTORY, COR_PRIMARIA, "compacto"),
            self._criar_card_stat("Hoje", stats['hoje'], ft.Icons.TODAY, COR_SUCESSO, "compacto"),
            self._criar_card_stat("Semana", stats['semana'], ft.Icons.DATE_RANGE, COR_SECUNDARIA, "compacto"),
            self._criar_card_stat("Mês", stats['mes'], ft.Icons.CALENDAR_MONTH, COR_ALERTA, "compacto"),
        ], spacing=10)

        # Carregar logs iniciais
        atualizar_lista_logs()

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.HISTORY, color=COR_PRIMARIA, size=28),
                ft.Text("Log do Sistema"),
            ]),
            content=ft.Container(
                content=ft.Column([
                    # Estatísticas
                    cards_stats,

                    ft.Container(height=10),

                    # Filtros
                    ft.Container(
                        content=ft.Row([
                            dropdown_categoria,
                            dropdown_tipo_acao,
                            campo_pesquisa,
                            ft.IconButton(
                                icon=ft.Icons.CLEAR,
                                tooltip="Limpar filtros",
                                on_click=limpar_filtros,
                                icon_color=COR_ERRO,
                            ),
                        ], spacing=10),
                        padding=10,
                        bgcolor=ft.Colors.GREY_100,
                        border_radius=8,
                    ),

                    ft.Container(height=5),

                    # Lista de logs
                    ft.Container(
                        content=lista_logs,
                        height=380,
                        border=ft.border.all(1, COR_CINZA_CLARO),
                        border_radius=8,
                    ),

                    # Paginação
                    ft.Container(
                        content=ft.Row([
                            btn_anterior,
                            texto_paginacao,
                            btn_proximo,
                        ], alignment=ft.MainAxisAlignment.CENTER, spacing=10),
                        padding=ft.padding.only(top=10),
                    ),

                ], spacing=5),
                width=750,
                height=580,
            ),
            actions=[
                ft.ElevatedButton("Fechar", on_click=fechar, bgcolor=COR_PRIMARIA, color="white"),
            ],
        )

        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()

    def _mostrar_sobre(self):
        """Exibe a tela Sobre com informações do desenvolvedor e do programa."""

        def fechar(ev):
            dialog.open = False
            self.page.update()

        # Informações do programa
        versao = "1.05"
        data_atualizacao = "Dezembro de 2025"
        desenvolvedor = "Fernando Albéniz Machado de Moura Guedes"
        contato = "(83) 9 9638-1689"

        # Descrição do programa
        descricao = """O Sistema de Gestão de RH é uma solução completa desenvolvida especialmente para a RENOVO Montagens Industriais, visando otimizar e automatizar os processos de gestão de recursos humanos da empresa.

Este sistema foi projetado para facilitar o gerenciamento de colaboradores, contratos, documentos e demais informações essenciais para o departamento de RH."""

        # Lista de funcionalidades
        funcionalidades = [
            ("Gestão de Colaboradores", "Cadastro completo com dados pessoais, profissionais e documentação", ft.Icons.PEOPLE),
            ("Sistema de Login", "Autenticação segura com níveis de acesso (Admin, Operador, Visualizador)", ft.Icons.LOGIN),
            ("Gerenciamento de Usuários", "Criação, edição e controle de acesso dos usuários do sistema", ft.Icons.ADMIN_PANEL_SETTINGS),
            ("Contratos de Experiência", "Acompanhamento de prazos, renovações e conversão automática para CLT", ft.Icons.SCHEDULE),
            ("Controle de Férias", "Monitoramento de períodos aquisitivos e concessivos", ft.Icons.BEACH_ACCESS),
            ("Gestão de Documentos", "Anexação e controle de documentos obrigatórios dos colaboradores", ft.Icons.FOLDER),
            ("Banco de Talentos", "Histórico completo de todos os colaboradores que passaram pela empresa", ft.Icons.FOLDER_SHARED),
            ("Block-List", "Registro de ex-colaboradores com histórico de desligamentos", ft.Icons.BLOCK),
            ("Gestão de Empresas", "Cadastro e gerenciamento de empresas contratantes", ft.Icons.BUSINESS),
            ("Exportação Excel", "Geração de relatórios em formato Excel", ft.Icons.DOWNLOAD),
            ("Backup de Dados", "Sistema de backup para proteção das informações", ft.Icons.BACKUP),
            ("Histórico de Alterações", "Registro de todas as modificações com identificação do usuário", ft.Icons.HISTORY),
            ("Geração de PDF", "Ficha de registro em formato PDF para impressão", ft.Icons.PICTURE_AS_PDF),
            ("Alertas Automáticos", "Notificações de contratos e férias vencendo", ft.Icons.NOTIFICATIONS),
            ("Recuperação de Senha", "Recuperação via pergunta de segurança", ft.Icons.LOCK_RESET),
            ("Localização", "Atribuição de localização ao colaborador com histórico de movimentações", ft.Icons.LOCATION_ON),
        ]

        # Criar widgets das funcionalidades
        lista_funcionalidades = []
        for titulo, desc, icone in funcionalidades:
            lista_funcionalidades.append(
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(icone, size=20, color="white"),
                            bgcolor=COR_PRIMARIA,
                            padding=8,
                            border_radius=8,
                        ),
                        ft.Column([
                            ft.Text(titulo, weight=ft.FontWeight.BOLD, size=13),
                            ft.Text(desc, size=11, color=ft.Colors.GREY_700),
                        ], spacing=2, expand=True),
                    ], spacing=12),
                    padding=ft.padding.symmetric(vertical=6),
                )
            )

        # Novidades da versão
        novidades = [
            "Correção de bugs no sistema de controle de férias",
            "Nova funcionalidade de Localização do colaborador com histórico completo",
        ]

        lista_novidades = []
        for novidade in novidades:
            lista_novidades.append(
                ft.Row([
                    ft.Icon(ft.Icons.CHECK_CIRCLE, size=16, color=COR_SUCESSO),
                    ft.Text(novidade, size=12),
                ], spacing=8)
            )

        # Caminho da logo
        logo_path = os.path.join(get_base_path(), "imagens", "Logomarca Renovo.png")

        # Widget da logo
        if os.path.exists(logo_path):
            logo_widget = ft.Container(
                content=ft.Image(src=logo_path, width=200, height=80, fit=ft.ImageFit.CONTAIN),
                bgcolor="#e8f4fc",
                padding=15,
                border_radius=10,
            )
        else:
            logo_widget = ft.Container(
                content=ft.Column([
                    ft.Icon(ft.Icons.BUSINESS, size=40, color=COR_PRIMARIA),
                    ft.Text("RENOVO", size=18, weight=ft.FontWeight.BOLD, color=COR_PRIMARIA),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=5),
                bgcolor="#e8f4fc",
                padding=15,
                border_radius=10,
            )

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ft.Icons.INFO, color=COR_PRIMARIA, size=28),
                ft.Text("Sobre o Sistema"),
            ]),
            content=ft.Container(
                content=ft.Column([
                    # Header com logo e informações básicas
                    ft.Container(
                        content=ft.Row([
                            logo_widget,
                            ft.Column([
                                ft.Text("Sistema de Gestão de RH", size=18, weight=ft.FontWeight.BOLD, color=COR_PRIMARIA),
                                ft.Text("RENOVO Montagens Industriais", size=13, color=ft.Colors.GREY_700),
                                ft.Container(height=10),
                                ft.Row([
                                    ft.Container(
                                        content=ft.Text(f"Versão {versao}", color="white", size=11),
                                        bgcolor=COR_SUCESSO,
                                        padding=ft.padding.symmetric(horizontal=10, vertical=4),
                                        border_radius=4,
                                    ),
                                    ft.Text(f"Atualizado em {data_atualizacao}", size=11, color=ft.Colors.GREY_600),
                                ], spacing=10),
                            ], spacing=3, expand=True),
                        ], spacing=20),
                        padding=15,
                        bgcolor=ft.Colors.GREY_100,
                        border_radius=10,
                    ),

                    ft.Container(height=15),

                    # Desenvolvedor
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ft.Icons.CODE, size=20, color=COR_SECUNDARIA),
                            ft.Column([
                                ft.Text("Desenvolvido por", size=11, color=ft.Colors.GREY_600),
                                ft.Text(desenvolvedor, weight=ft.FontWeight.BOLD, size=14),
                                ft.Row([
                                    ft.Icon(ft.Icons.PHONE, size=14, color=ft.Colors.GREY_600),
                                    ft.Text(contato, size=12, color=ft.Colors.GREY_700),
                                ], spacing=5),
                            ], spacing=2),
                        ], spacing=12),
                        padding=12,
                        border=ft.border.all(1, COR_SECUNDARIA),
                        border_radius=8,
                    ),

                    ft.Container(height=15),

                    # Novidades da versão
                    ft.Text(f"Novidades da Versão {versao}", weight=ft.FontWeight.BOLD, size=14, color=COR_SUCESSO),
                    ft.Container(
                        content=ft.Column(lista_novidades, spacing=5),
                        padding=10,
                        bgcolor="#e8f5e9",
                        border_radius=8,
                    ),

                    ft.Container(height=15),

                    # Funcionalidades
                    ft.Text("Funcionalidades", weight=ft.FontWeight.BOLD, size=14, color=COR_PRIMARIA),
                    ft.Container(
                        content=ft.Column(lista_funcionalidades, scroll=ft.ScrollMode.AUTO, spacing=0),
                        height=180,
                        padding=10,
                        border=ft.border.all(1, COR_CINZA_CLARO),
                        border_radius=8,
                    ),

                    ft.Container(height=10),

                    # Footer
                    ft.Container(
                        content=ft.Text(
                            f"© {datetime.now().year} RENOVO Montagens Industriais. Todos os direitos reservados.",
                            size=10,
                            color=ft.Colors.GREY_500,
                            text_align=ft.TextAlign.CENTER,
                        ),
                        alignment=ft.alignment.center,
                    ),
                ], scroll=ft.ScrollMode.AUTO, spacing=0),
                width=600,
                height=580,
            ),
            actions=[
                ft.ElevatedButton("Fechar", on_click=fechar, bgcolor=COR_PRIMARIA, color="white"),
            ],
        )

        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()

    def _voltar_ao_erp(self, e):
        """Fecha o módulo e sinaliza para o ERP restaurar da bandeja"""
        if ERP_USER_DATA.erp_path:
            # Registrar log de saída
            usuario = db.get_usuario_logado()
            if usuario:
                db.registrar_log(
                    tipo_acao='SISTEMA',
                    categoria='APLICACAO',
                    descricao='Retornando ao ERP',
                    usuario=usuario.get('nome_completo')
                )

            # Criar arquivo de sinal para o ERP restaurar da bandeja
            erp_dir = os.path.dirname(ERP_USER_DATA.erp_path)
            signal_file = os.path.join(erp_dir, ".erp_restore_signal")
            try:
                with open(signal_file, "w") as f:
                    f.write("restore")
            except Exception:
                pass

            # Fechar o módulo
            self.page.window.close()


def main(page: ft.Page):
    """Função principal."""
    # Se veio do ERP, definir um usuário padrão para o sistema funcionar
    if ERP_USER_DATA.nome:
        # Criar usuário temporário baseado nos dados do ERP
        usuario_erp = {
            'id': 0,
            'login': ERP_USER_DATA.usuario or 'erp_user',
            'nome_completo': ERP_USER_DATA.nome,
            'nivel_acesso': 'administrador',  # Acesso total quando vem do ERP
            'cargo': ERP_USER_DATA.cargo or 'Usuário ERP'
        }
        db.set_usuario_logado(usuario_erp)
    else:
        # Executando diretamente (sem ERP) - definir usuário administrador padrão
        usuario_padrao = {
            'id': 0,
            'login': 'admin',
            'nome_completo': 'Administrador',
            'nivel_acesso': 'administrador',
            'cargo': 'Administrador do Sistema'
        }
        db.set_usuario_logado(usuario_padrao)

    # Ir direto para o sistema (sem splash e sem login)
    SistemaRH(page)


if __name__ == "__main__":
    # Usar janela desktop nativa (funciona com flet pack)
    ft.app(
        target=main,
        assets_dir="imagens",
        view=ft.AppView.FLET_APP
    )
